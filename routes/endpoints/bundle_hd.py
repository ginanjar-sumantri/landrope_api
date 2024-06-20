from uuid import UUID, uuid4
from fastapi import APIRouter, status, Depends, HTTPException
from fastapi.responses import StreamingResponse
from fastapi_pagination import Params
from fastapi_async_sqlalchemy import db
from sqlmodel import select, and_, or_, update
from models.bundle_model import BundleHd
from models.worker_model import Worker
from models.bidang_model import Bidang
from models.kjb_model import KjbHd, KjbDt
from models.planing_model import Planing
from models.checklist_kelengkapan_dokumen_model import ChecklistKelengkapanDokumenDt
from schemas.bundle_hd_sch import (BundleHdSch, BundleHdCreateSch, BundleHdUpdateSch, BundleHdByIdSch)
from schemas.bundle_dt_sch import BundleDtCreateSch
from schemas.bidang_sch import BidangUpdateSch
from schemas.response_sch import (PostResponseBaseSch, GetResponseBaseSch, GetResponsePaginatedSch, PutResponseBaseSch, create_response)
from common.exceptions import (IdNotFoundException, ImportFailedException)
from common.generator import generate_code
from common.enum import JenisBayarEnum
from models.code_counter_model import CodeCounterEnum
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from shapely import wkt, wkb
from itertools import islice
import crud
import json


router = APIRouter()

@router.post("/create", response_model=PostResponseBaseSch[BundleHdSch], status_code=status.HTTP_201_CREATED)
async def create(sch: BundleHdCreateSch,
                current_worker: Worker = Depends(crud.worker.get_active_worker)):
    
    """Create a new object"""
        
    new_obj = await crud.bundlehd.create_and_generate(obj_in=sch, created_by_id=current_worker.id)
    new_obj = await crud.bundlehd.get_by_id(id=new_obj.id)
    
    return create_response(data=new_obj)

@router.get("", response_model=GetResponsePaginatedSch[BundleHdSch])
async def get_list(
            params: Params=Depends(), 
            order_by:str = None, 
            keyword:str = None, 
            filter_query:str = None,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(BundleHd
                            ).outerjoin(Bidang, Bidang.bundle_hd_id == BundleHd.id
                            ).outerjoin(KjbDt, KjbDt.bundle_hd_id == BundleHd.id
                            ).outerjoin(Planing, Planing.id == BundleHd.planing_id)
    
    if keyword:
        query = query.filter(
            or_(
                BundleHd.code.ilike(f'%{keyword}%'),
                Planing.name.ilike(f'%{keyword}%'),
                KjbDt.alashak.ilike(f'%{keyword}%'),
                Bidang.id_bidang.ilike(f'%{keyword}%')
            ))
    
    if filter_query:
        filter_query = json.loads(filter_query)
        for key, value in filter_query.items():
            query = query.where(getattr(BundleHd, key) == value)
    
    query = query.distinct()

    objs = await crud.bundlehd.get_multi_paginated_ordered(params=params, query=query, order_by="updated_at")

    return create_response(data=objs)

@router.get("/{id}", response_model=GetResponseBaseSch[BundleHdByIdSch])
async def get_by_id(id:UUID):

    """Get an object by id"""

    obj = await crud.bundlehd.get_by_id(id=id)
    if obj:
        return create_response(data=obj)
    else:
        raise IdNotFoundException(BundleHd, id)

@router.put("/{id}", response_model=PutResponseBaseSch[BundleHdSch])
async def update_(id:UUID, 
                 sch:BundleHdUpdateSch,
                 current_worker: Worker = Depends(crud.worker.get_active_worker)):
    
    """Update a obj by its id"""

    obj_current = await crud.bundlehd.get(id=id)

    if not obj_current:
        raise IdNotFoundException(BundleHd, id)
    
    obj_updated = await crud.bundlehd.update(obj_current=obj_current, obj_new=sch, updated_by_id=current_worker.id)
    obj_updated = await crud.bundlehd.get_by_id(id=obj_updated.id)
    return create_response(data=obj_updated)

@router.put("regenerate/{id}", response_model=PutResponseBaseSch[BundleHdSch])
async def regenerate(id:UUID,
                     current_worker: Worker = Depends(crud.worker.get_active_worker)):

    """Get an object by id"""

    obj = await crud.bundlehd.get(id=id)
    if obj is None:
        raise IdNotFoundException(BundleHd, id)
    
    db_session = db.session

    bundle_dts = await crud.bundledt.get_by_bundle_hd_id(bundle_hd_id=obj.id)
    list_ids = [dt.dokumen_id for dt in bundle_dts]
    dokumens = await crud.dokumen.get_not_in_by_ids(list_ids=list_ids)
    
    if dokumens:
        for doc in dokumens:
            if doc.is_active == False:
                continue
            code = obj.code + doc.code
            bundle_dt_create = BundleDtCreateSch(code=code, 
                                        dokumen_id=doc.id,
                                        bundle_hd_id=obj.id)
            
            await crud.bundledt.create(obj_in=bundle_dt_create, db_session=db_session, with_commit=False)
    
    obj_new = obj

    obj_updated = await crud.bundlehd.update(obj_current=obj, 
                                             obj_new=obj_new, 
                                             db_session=db_session, 
                                             updated_by_id=current_worker.id,
                                             with_commit=True)
    
    obj_updated = await crud.bundlehd.get_by_id(id=obj_updated.id)
    
    return create_response(data=obj_updated)

@router.get("/generate/bundle/hd")
async def generate_bundle(start:int, size:int):

    """Get an object by id"""

    db_session = db.session
    notes = []

    db_session_a = db.session

    query = "select * from import_id_bidang_lama"

    response = await db_session_a.execute(query)
    result = response.all()
    counter:int = 1

    for id_bidang in islice(result, start, None):
        if counter == size:
            break
        
        id_bidang_lama = id_bidang.id_bidang_lama
        if id_bidang_lama is None:
            counter += 1
            continue
        
        bidang = await crud.bidang.get_by_id_bidang_lama(idbidang_lama=id_bidang_lama.replace(' ', ''))
        if bidang is None:
            counter += 1
            bidang_not_found = {"id_bidang" : id_bidang_lama, "note" : "Bidang Not Found"}
            notes.append(bidang_not_found)
            continue

        if bidang.bundle_hd_id:
            bidang_has_bundle = {"id_bidang" : id_bidang_lama, "note" : "Bidang Has Bundle"}
            notes.append(bidang_has_bundle)
            counter += 1
            continue

        bundle_sch = BundleHdCreateSch(planing_id=bidang.planing_id, keyword=bidang.alashak)
        bundle = await crud.bundlehd.create_and_generate(obj_in=bundle_sch, db_session=db_session, with_commit=False)

        update_query = update(Bidang).where(Bidang.id == bidang.id).values(bundle_hd_id=bundle.id)
        await db_session.execute(update_query)

        bidang_has_create_bundle = {"id_bidang" : id_bidang_lama, "note" : "Bidang Has Create Bundle"}
        notes.append(bidang_has_create_bundle)

        counter += 1
        
    
    wb = Workbook()
    ws = wb.active

    header_string = ["Id Bidang", "Note"]

    for idx, val in enumerate(header_string):
        ws.cell(row=1, column=idx + 1, value=val).font = Font(bold=True)
    
    x = 1
    for row_data in notes:
        x += 1
        ws.cell(row=x, column=1, value=row_data["id_bidang"])
        ws.cell(row=x, column=2, value=row_data["note"])
    
    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)

    await db_session.commit()
    
    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=Status Bidang Generate Bundle.xlsx"})

@router.put("regenerate/bundle/kelengkapan")
async def regenerate(dokumen_name:str):

    """Get an object by id"""
    size = 1000
    db_session = db.session

    dokumen = await crud.dokumen.get_by_name(name=dokumen_name)
    if dokumen is None:
        raise HTTPException(status_code=422, detail=f"Dokumen {dokumen_name} not found in master!")


    bundle_hd_ids = await crud.bundlehd.get_by_dokumen_not_exists(dokumen_id=dokumen.id)
    data = 1
    for bundle_hd_id in bundle_hd_ids:
        bundle_hd = await crud.bundlehd.get(id=bundle_hd_id.id)
        if bundle_hd is None:
            continue
        bundle_dt = await crud.bundledt.get_by_bundle_hd_id_and_dokumen_id(bundle_hd_id=bundle_hd.id, dokumen_id=dokumen.id)
        if bundle_dt is None:
            code = bundle_hd.code + dokumen.code
            bundle_dt_new = BundleDtCreateSch(code=code, dokumen_id=dokumen.id, bundle_hd_id=bundle_hd.id)
            bundle_dt = await crud.bundledt.create(obj_in=bundle_dt_new, db_session=db_session, with_commit=False)

        bidang = await crud.bidang.get_by_bundle_hd_id(bundle_hd_id=bundle_hd.id)
        if bidang:
            kelengkapan_dokumen_header = await crud.checklist_kelengkapan_dokumen_hd.get_by_bidang_id(bidang_id=bidang.id)
            kelengkapan_dokumen_detail = await crud.checklist_kelengkapan_dokumen_dt.get_by_checklist_kelengkapan_dokumen_hd_id_and_dokumen_id(checklist_kelengkapan_dokumen_hd_id=kelengkapan_dokumen_header.id, dokumen_id=dokumen.id)
            if kelengkapan_dokumen_detail is None:
                kelengkapan_dokumen_detail_new = ChecklistKelengkapanDokumenDt(checklist_kelengkapan_dokumen_hd_id=kelengkapan_dokumen_header.id,
                                                                               bundle_dt_id=bundle_dt.id,
                                                                               jenis_bayar=JenisBayarEnum.DP,
                                                                               dokumen_id=dokumen.id
                                                                               )
                kelengkapan_dokumen_detail_new.id = uuid4()
                await crud.checklist_kelengkapan_dokumen_dt.create(obj_in=kelengkapan_dokumen_detail_new, db_session=db_session, with_commit=False)
        data += 1
        if data == size:
            break
        
    return create_response({"detail": "SUCCESS"})