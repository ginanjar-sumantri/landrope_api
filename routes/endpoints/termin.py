from uuid import UUID, uuid4
from fastapi import APIRouter, status, Depends, Request, HTTPException, Response, UploadFile, BackgroundTasks
from fastapi.responses import StreamingResponse
from fastapi_pagination import Params
from fastapi_async_sqlalchemy import db
from sqlmodel import select, or_, and_, func
from sqlalchemy import cast, String, Date
from sqlalchemy.orm import selectinload
import crud
from models import (Termin, Worker, Invoice, InvoiceDetail, Tahap, TahapDetail, KjbHd, Spk, Bidang, PaymentDetail, Payment, PaymentGiroDetail, 
                    Planing, Workflow, WorkflowNextApprover, BidangKomponenBiaya, Planing, Project, Desa, Ptsk)
from schemas.tahap_sch import TahapSrcSch
from schemas.termin_sch import (TerminSch, TerminCreateSch, TerminUpdateSch, TerminByIdSch, TerminBidangIDSch, TerminIdSch, TerminVoidSch, TerminFilterJson, TerminUpdateBaseSch)
from schemas.invoice_sch import (InvoiceHistoryInTermin)
from schemas.spk_sch import SpkSrcSch, SpkInTerminSch, SpkIdSch, SpkJenisBayarSch
from schemas.kjb_hd_sch import KjbHdForTerminByIdSch, KjbHdSearchSch
from schemas.bidang_sch import BidangForUtjSch
from schemas.bidang_komponen_biaya_sch import BidangKomponenBiayaSch
from schemas.marketing_sch import ManagerSrcSch, SalesSrcSch
from schemas.rekening_sch import RekeningSch
from schemas.notaris_sch import NotarisSrcSch
from schemas.beban_biaya_sch import BebanBiayaEstimatedAmountSch
from schemas.workflow_sch import WorkflowSystemCreateSch, WorkflowSystemAttachmentSch, WorkflowUpdateSch
from schemas.response_sch import (GetResponseBaseSch, GetResponsePaginatedSch, 
                                  PostResponseBaseSch, PutResponseBaseSch, create_response)
from common.exceptions import (IdNotFoundException, DocumentFileNotFoundException)
from common.ordered import OrderEnumSch
from common.enum import (JenisBayarEnum, WorkflowEntityEnum, WorkflowLastStatusEnum, SatuanBayarEnum, SatuanHargaEnum,
                        jenis_bayar_to_termin_status_pembebasan_dict, jenis_bayar_to_code_counter_enum, jenis_bayar_to_text)
from services.gcloud_storage_service import GCStorageService
from services.helper_service import KomponenBiayaHelper
from services.workflow_service import WorkflowService
from services.rfp_service import RfpService
from services.termin_service import TerminService
from services.encrypt_service import encrypt_id

from decimal import Decimal

from io import BytesIO
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font
import json

import time

router = APIRouter()

@router.get("", response_model=GetResponsePaginatedSch[TerminSch])
async def get_list(
            params: Params=Depends(), 
            order_by:str = None, 
            keyword:str = None,
            is_utj:bool = False,
            filter_query:str | None = None,
            filter_list: str | None = None,
            filter_json: str | None = None,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    jenis_bayars = []
    if is_utj:
        jenis_bayars = [JenisBayarEnum.UTJ.value, 
                        JenisBayarEnum.UTJ_KHUSUS.value]
    else:
        jenis_bayars = [JenisBayarEnum.DP.value, 
                        JenisBayarEnum.LUNAS.value, 
                        JenisBayarEnum.PENGEMBALIAN_BEBAN_PENJUAL.value, 
                        JenisBayarEnum.BEGINNING_BALANCE.value, 
                        JenisBayarEnum.BIAYA_LAIN.value, 
                        JenisBayarEnum.SISA_PELUNASAN.value,
                        JenisBayarEnum.PELUNASAN.value]

    query = select(Termin).outerjoin(Invoice, Invoice.termin_id == Termin.id
                        ).outerjoin(Tahap, Tahap.id == Termin.tahap_id
                        ).outerjoin(KjbHd, KjbHd.id == Termin.kjb_hd_id
                        ).outerjoin(Spk, Spk.id == Invoice.spk_id
                        ).outerjoin(Bidang, Bidang.id == Invoice.bidang_id
                        ).outerjoin(Ptsk, Ptsk.id == Tahap.ptsk_id
                        ).outerjoin(Planing, Planing.id == Tahap.planing_id
                        ).outerjoin(Project, Project.id == Planing.project_id
                        ).outerjoin(Desa, Desa.id == Planing.desa_id
                        ).outerjoin(Workflow, and_(Workflow.reference_id == Termin.id, Workflow.entity == WorkflowEntityEnum.TERMIN)
                        ).where(Termin.jenis_bayar.in_(jenis_bayars)).distinct()
    
    if filter_list == "list_approval":
        subquery_workflow = (select(Workflow.reference_id).join(Workflow.workflow_next_approvers
                                ).where(and_(WorkflowNextApprover.email == current_worker.email, 
                                            Workflow.entity == WorkflowEntityEnum.TERMIN,
                                            Workflow.last_status != WorkflowLastStatusEnum.COMPLETED))
                    ).distinct()
        
        query = query.filter(Termin.id.in_(subquery_workflow))
    
    if keyword and keyword != '':
        query = query.filter(
            or_(
                Termin.code.ilike(f'%{keyword}%'),
                Termin.jenis_bayar.ilike(f'%{keyword}%'),
                Termin.nomor_memo.ilike(f'%{keyword}%'),
                cast(Tahap.nomor_tahap, String).ilike(f'%{keyword}%'),
                KjbHd.code.ilike(f'%{keyword}%'),
                Bidang.id_bidang.ilike(f'%{keyword}%'),
                Bidang.alashak.ilike(f'%{keyword}%'),
                Ptsk.name.ilike(f'%{keyword}%'),
                Project.name.ilike(f'%{keyword}%'),
                Desa.name.ilike(f'%{keyword}%'),
                Tahap.group.ilike(f'%{keyword}%'),
                Workflow.last_status.ilike(f'%{keyword}%'),
                Workflow.step_name.ilike(f'%{keyword}%')
            )
        )

    if filter_json:
        json_loads = json.loads(filter_json)
        termin_filter_json = TerminFilterJson(**dict(json_loads))

        if termin_filter_json.code:
            query = query.filter(cast(Termin.code, String).ilike(f'%{termin_filter_json.code}%'))
        if termin_filter_json.project:
            query = query.filter(cast(Project.name, String).ilike(f'%{termin_filter_json.project}%'))
        if termin_filter_json.desa:
            query = query.filter(cast(Desa.name, String).ilike(f'%{termin_filter_json.desa}%'))
        if termin_filter_json.ptsk:
            query = query.filter(cast(Ptsk.name, String).ilike(f'%{termin_filter_json.ptsk}%'))
        if termin_filter_json.nomor_tahap:
            query = query.filter(cast(Tahap.nomor_tahap, String).ilike(f'%{termin_filter_json.nomor_tahap}%'))
        if termin_filter_json.group:
            query = query.filter(cast(Tahap.group, String).ilike(f'%{termin_filter_json.group}%'))
        if termin_filter_json.jenis_bayar:
            query = query.filter(cast(func.replace(Termin.jenis_bayar, '_', ' '), String).ilike(f'%{termin_filter_json.jenis_bayar}%'))
        if termin_filter_json.tanggal_pengajuan:
            query = query.filter(cast(cast(Termin.created_at, Date), String).ilike(f'%{termin_filter_json.tanggal_pengajuan}%'))
        if termin_filter_json.nomor_memo:
            query = query.filter(cast(Termin.nomor_memo, String).ilike(f'%{termin_filter_json.nomor_memo}%'))
        if termin_filter_json.status:
            query = query.filter(or_(cast(Workflow.last_status, String).ilike(f'%{termin_filter_json.status}%'), cast(Workflow.step_name, String).ilike(f'%{termin_filter_json.status}%')))
        if termin_filter_json.last_update_by:
            query = query.filter(cast(Worker.name, String).ilike(f'%{termin_filter_json.last_update_by}%'))

    if filter_query:
        filter_query = json.loads(filter_query)
        for key, value in filter_query.items():
                query = query.where(getattr(Termin, key) == value)
    
    query = query.distinct()

    objs = await crud.termin.get_multi_paginated_ordered(params=params, order_by="created_at", order=OrderEnumSch.descendent, query=query)

    items = []
    reference_ids = [termin.id for termin in objs.data.items]
    workflows = await crud.workflow.get_by_reference_ids(reference_ids=reference_ids, entity=WorkflowEntityEnum.TERMIN)

    for obj in objs.data.items:
        workflow = next((wf for wf in workflows if wf.reference_id == obj.id), None)
        if workflow:
            obj.status_workflow = workflow.last_status
            obj.step_name_workflow = workflow.step_name

        items.append(obj)

    objs.data.items = items

    return create_response(data=objs)

@router.get("/{id}", response_model=GetResponseBaseSch[TerminByIdSch])
async def get_by_id(id:UUID,
                    current_worker:Worker = Depends(crud.worker.get_active_worker)):

    """Get an object by id"""
    
    obj = await crud.termin.get_by_id(id=id)
    if obj:
        return create_response(data=obj)
    else:
        raise IdNotFoundException(Termin, id)

@router.post("/create", response_model=PostResponseBaseSch[TerminSch], status_code=status.HTTP_201_CREATED)
async def create(
            sch: TerminCreateSch,
            request: Request,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Create a new object"""

    db_session = db.session
    sch.is_void = False

    if (sch.is_draft or False) is False:
        if sch.jenis_bayar not in [JenisBayarEnum.UTJ, JenisBayarEnum.UTJ_KHUSUS]:
            await TerminService().filter_termin(sch=sch)

    new_obj = await TerminService().create_termin(sch=sch, db_session=db_session, current_worker=current_worker, request=request)
    return_obj = await crud.termin.get_by_id(id=new_obj.id)

    return create_response(data=return_obj)

@router.put("/{id}", response_model=PutResponseBaseSch[TerminSch])
async def update_(
            id:UUID,
            sch:TerminUpdateSch,
            request:Request,
            background_task:BackgroundTasks,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Update a obj by its id"""

    db_session = db.session
    db_session.autoflush = False

    obj_current = await crud.termin.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Termin, id)
    
    if obj_current.is_void:
        raise HTTPException(status_code=422, detail="Memo Sudah divoid")
    
    last_status_current:WorkflowLastStatusEnum = WorkflowLastStatusEnum.DRAFT
    
    workflow = await crud.workflow.get_by_reference_id(reference_id=obj_current.id)
    if workflow:
        last_status_current = workflow.last_status

    # JIKA HANYA UPDATE NOMOR MEMO DAN UPLOAD FILE MEMO BAYAR
    if last_status_current == WorkflowLastStatusEnum.NEED_DATA_UPDATE:
        object_update_memo = await TerminService().update_nomor_memo_dan_file(sch=sch, obj_current=obj_current, worker_id=current_worker.id, request=request)
        object_update_memo = await crud.termin.get_by_id(id=object_update_memo.id)

        if object_update_memo.jenis_bayar not in [JenisBayarEnum.UTJ_KHUSUS, JenisBayarEnum.UTJ]:
            background_task.add_task(TerminService.generate_printout_memo_bayar, object_update_memo.id)
            background_task.add_task(TerminService.merge_memo_signed, object_update_memo.id)

        return create_response(data=object_update_memo)

    else:
        if (sch.is_draft or False) is False:
            if sch.jenis_bayar not in [JenisBayarEnum.UTJ_KHUSUS, JenisBayarEnum.UTJ]:
                msg_error_wf = "Memo Approval Has Been Completed!" if WorkflowLastStatusEnum.COMPLETED else "Memo Approval Need Approval!"

                if last_status_current not in [WorkflowLastStatusEnum.NEED_DATA_UPDATE, WorkflowLastStatusEnum.REJECTED, WorkflowLastStatusEnum.DRAFT]:
                    raise HTTPException(status_code=422, detail=f"Failed update. Detail : {msg_error_wf}")
                
                if last_status_current in [WorkflowLastStatusEnum.NEED_DATA_UPDATE, WorkflowLastStatusEnum.DRAFT, WorkflowLastStatusEnum.REJECTED]:
                    await TerminService().filter_termin(sch=sch)

        if sch.jenis_bayar in [JenisBayarEnum.UTJ_KHUSUS, JenisBayarEnum.UTJ]:
            kjb_hd_current = await crud.kjb_hd.get(id=sch.kjb_hd_id)
            total_utj = sum([dt.amount for dt in sch.invoices])

            if total_utj > kjb_hd_current.utj_amount:
                raise HTTPException(status_code=422, detail="Total UTJ tidak boleh lebih besar dari UTJ amount di KJB")
            
        obj_updated = await TerminService().edit_termin(obj_current=obj_current, sch=sch, db_session=db_session, current_worker=current_worker, request=request)

        if obj_updated.jenis_bayar not in [JenisBayarEnum.UTJ_KHUSUS, JenisBayarEnum.UTJ]:
            background_task.add_task(TerminService.generate_printout_memo_bayar, obj_updated.id)
            background_task.add_task(TerminService.merge_memo_signed, obj_updated.id)
        else:
            background_task.add_task(TerminService.generate_printout_utj, obj_updated.id)

        obj_updated = await crud.termin.get_by_id(id=obj_updated.id)

        return create_response(data=obj_updated)

# VOID MEMO BAYAR
@router.put("/void/{id}", response_model=GetResponseBaseSch[TerminByIdSch])
async def void(id:UUID, 
            sch:TerminVoidSch,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """void a obj by its ids"""
    db_session = db.session

    obj_current = await crud.termin.get_by_id(id=id)

    if not obj_current:
        raise IdNotFoundException(Termin, id)
    
    obj_updated = await TerminService().void(sch=sch, obj_current=obj_current, db_session=db_session, current_worker=current_worker)

    obj_updated = await crud.termin.get_by_id(id=obj_updated.id)
    
    return create_response(data=obj_updated) 

# DOWNLOAD FILE MEMO BAYAR SIGNED
@router.get("/download-file/{id}")
async def download_file(id:UUID):
    """Download File Dokumen"""

    obj_current = await crud.termin.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Termin, id)
    if obj_current.file_upload_path is None:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=obj_current.file_upload_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    
    ext = obj_current.file_path.split('.')[-1]

    # return FileResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={obj_current.id}.{ext}"})
    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename=Hasil Peta Lokasi-{id}-{obj_current.code}.{ext}"
    return response

@router.get("/estimated/amount", response_model=GetResponseBaseSch[BebanBiayaEstimatedAmountSch])
async def get_estimated_amount(bidang_id:UUID, beban_biaya_id:UUID):

    master_beban_biaya = await crud.bebanbiaya.get(id=beban_biaya_id)
    if master_beban_biaya is None:
        raise HTTPException(status_code=422, detail="master beban biaya tidak ditemukan")
    
    result = BebanBiayaEstimatedAmountSch.from_orm(master_beban_biaya)

    bidang_komponen_biaya_current = await crud.bidang_komponen_biaya.get_by_bidang_id_and_beban_biaya_id(bidang_id=bidang_id, beban_biaya_id=beban_biaya_id)

    if master_beban_biaya.satuan_bayar == SatuanBayarEnum.Amount and master_beban_biaya.satuan_harga == SatuanHargaEnum.Lumpsum:
        if bidang_komponen_biaya_current:
            result.estimated_amount = bidang_komponen_biaya_current.estimated_amount
            result.bidang_id = bidang_id
            result.amount = bidang_komponen_biaya_current.amount
            result.beban_biaya_id = master_beban_biaya.id
        else:
            result.estimated_amount = master_beban_biaya.amount 
            result.bidang_id = bidang_id
            result.amount = master_beban_biaya.amount
            result.beban_biaya_id = master_beban_biaya.id
    else:
        if bidang_komponen_biaya_current:
            result.estimated_amount = bidang_komponen_biaya_current.estimated_amount
            result.bidang_id = bidang_id
            result.beban_biaya_id = master_beban_biaya.id
        else:
            estimated_amount = await KomponenBiayaHelper().get_estimated_amount_v2(bidang_id=bidang_id, formula=master_beban_biaya.formula, beban_biaya_id=master_beban_biaya.id)
            result.estimated_amount = estimated_amount
            result.bidang_id = bidang_id
            result.beban_biaya_id = master_beban_biaya.id

    return create_response(data=result)

@router.get("/estimated/amount/edited", response_model=GetResponseBaseSch[BebanBiayaEstimatedAmountSch])
async def get_estimated_amount_edited(bidang_id:UUID, beban_biaya_id:UUID, 
                                    amount:Decimal | None = None, 
                                    satuan_bayar:SatuanBayarEnum | None = None, 
                                    satuan_harga:SatuanHargaEnum | None = None):

    master_beban_biaya = await crud.bebanbiaya.get(id=beban_biaya_id)
    if master_beban_biaya is None:
        raise HTTPException(status_code=422, detail="master beban biaya tidak ditemukan")
    
    result = BebanBiayaEstimatedAmountSch.from_orm(master_beban_biaya)

    bidang_komponen_biaya_current = await crud.bidang_komponen_biaya.get_by_bidang_id_and_beban_biaya_id(bidang_id=bidang_id, beban_biaya_id=beban_biaya_id)

    if satuan_bayar == SatuanBayarEnum.Amount and satuan_harga == SatuanHargaEnum.Lumpsum:
        if bidang_komponen_biaya_current:
            result.estimated_amount = amount
            result.bidang_id = bidang_id
            result.amount = amount
            result.beban_biaya_id = master_beban_biaya.id
        else:
            result.estimated_amount = amount 
            result.bidang_id = bidang_id
            result.amount = amount
            result.beban_biaya_id = master_beban_biaya.id
    else:
        if bidang_komponen_biaya_current:
            if amount:
                estimated_amount = await KomponenBiayaHelper().get_estimated_amount_v2(bidang_id=bidang_id, formula=master_beban_biaya.formula, beban_biaya_id=master_beban_biaya.id, amount=amount)
                estimated_amount = estimated_amount
                result.estimated_amount = estimated_amount
                result.bidang_id = bidang_id
                result.amount = amount
                result.beban_biaya_id = master_beban_biaya.id
            else:
                result.estimated_amount = bidang_komponen_biaya_current.estimated_amount
                result.bidang_id = bidang_id
                result.beban_biaya_id = master_beban_biaya.id
        else:
            estimated_amount = await KomponenBiayaHelper().get_estimated_amount_v2(bidang_id=bidang_id, formula=master_beban_biaya.formula, beban_biaya_id=master_beban_biaya.id, amount=amount)
            result.estimated_amount = estimated_amount
            result.bidang_id = bidang_id
            result.beban_biaya_id = master_beban_biaya.id

    return create_response(data=result)

# PRINTOUT MEMO BAYAR
@router.get("/print-out/{id}")
async def printout(id:UUID | str, current_worker:Worker = Depends(crud.worker.get_active_worker)):

    """Print out Memo Bayar"""
    obj_current = await crud.termin.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Termin, id)
    
    workflow = await crud.workflow.get_by_reference_id(reference_id=obj_current.id)
    
    if workflow.last_status != WorkflowLastStatusEnum.COMPLETED:
        file_path = await TerminService().generate_printout_memo_bayar(id=id)
    else:
        file_path = obj_current.file_path

    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=file_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    
    ext = obj_current.file_path.split('.')[-1]

    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename={obj_current.code.replace('/', '_')}.{ext}"
    return response

# PRINTOUT UTJ
@router.get("/print-out/utj/{id}")
async def printout(id:UUID | str, current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Print out UTJ"""
    obj_current = await crud.termin.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Termin, id)
    
    
    file_path = await TerminService().generate_printout_utj(id=id)
    
    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=file_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    
    ext = obj_current.file_path.split('.')[-1]

    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename={obj_current.code.replace('/', '_')}.{ext}"
    return response

@router.get("/print-out/excel/{id}")
async def printout(id:UUID | str,
                   current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    obj_current = await crud.termin.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Termin, id)
    
    excel = await TerminService().generate_printout_to_excel(id=id)

    return StreamingResponse(excel, 
                                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                headers={"Content-Disposition": "attachment;filename=memo_pembayaran.xlsx"})  

# REPORT EXCEL MEMO BAYAR
@router.post("/export/excel")
async def get_report(
                termin_ids:TerminIdSch|None = None, 
                current_worker:Worker = Depends(crud.worker.get_active_worker)):

    filename:str = ''
    query = select(Termin)
    if termin_ids.termin_ids:
        query = query.filter(Termin.id.in_(termin_ids.termin_ids))
    
    query = query.filter(~Termin.jenis_bayar.in_([JenisBayarEnum.UTJ.value, JenisBayarEnum.UTJ_KHUSUS.value]))

    query = query.distinct()
    query = query.options(selectinload(Termin.invoices
                                ).options(selectinload(Invoice.bidang
                                                    ).options(selectinload(Bidang.pemilik)
                                                    ).options(selectinload(Bidang.planing
                                                                ).options(selectinload(Planing.project)
                                                                ).options(selectinload(Planing.desa))
                                                    ).options(selectinload(Bidang.invoices
                                                                ).options(selectinload(Invoice.payment_details)
                                                                ).options(selectinload(Invoice.termin))
                                                    )
                                ).options(selectinload(Invoice.payment_details
                                                    ).options(selectinload(PaymentDetail.payment
                                                                ).options(selectinload(Payment.giro))
                                                    ).options(selectinload(PaymentDetail.payment_giro
                                                                ).options(selectinload(PaymentGiroDetail.giro))
                                                    )
                                ).options(selectinload(Invoice.termin
                                                    ).options(selectinload(Termin.tahap)
                                                    )
                                ).options(selectinload(Invoice.details
                                                    ).options(selectinload(InvoiceDetail.bidang_komponen_biaya
                                                                        ).options(selectinload(BidangKomponenBiaya.beban_biaya)
                                                                        ).options(selectinload(BidangKomponenBiaya.bidang))
                                                    )
                                )
                )

    objs = await crud.termin.get_multi_no_page(query=query)

    data = []

    reference_ids = [s.id for s in objs]
    workflows = await crud.workflow.get_by_reference_ids(reference_ids=reference_ids, entity=WorkflowEntityEnum.TERMIN)

    wb = Workbook()
    ws = wb.active

    header_string = ["Id Bidang", 
                    "Id Bidang Lama",
                    "Project", 
                    "Desa",
                    "Nomor Tahap",
                    "Nomor Memo",
                    "Group",
                    "Pemilik",
                    "Alashak",
                    "Luas Surat",
                    "Luas Bayar", 
                    "Jumlah",
                    "Jenis Bayar",
                    "Status Workflow",
                    "Tanggal Last Workflow",
                    "Harga Transaksi", 
                    "Nomor Giro",
                    "Tanggal Pembayaran",
                    "Payment Status",
                    "Tanggal Last Payment Status"]

    for idx, val in enumerate(header_string):
        ws.cell(row=1, column=idx + 1, value=val).font = Font(bold=True)

    x = 1
    for termin in objs:
       for invoice in termin.invoices:
            workflow = next((wf for wf in workflows if wf.reference_id == termin.id), None)
            status_workflow = "-"
            last_status_workflow_at = None
            if workflow:
                status_workflow = workflow.last_status if workflow.last_status == WorkflowLastStatusEnum.COMPLETED else workflow.step_name or "-"
                last_status_workflow_at = workflow.last_status_at

            x += 1
            ws.cell(row=x, column=1, value=invoice.id_bidang)
            ws.cell(row=x, column=2, value=invoice.id_bidang_lama or "")
            ws.cell(row=x, column=3, value=invoice.bidang.project_name)
            ws.cell(row=x, column=4, value=invoice.bidang.desa_name)
            ws.cell(row=x, column=5, value=invoice.termin.nomor_tahap)
            ws.cell(row=x, column=6, value=invoice.termin.nomor_memo)
            ws.cell(row=x, column=7, value=invoice.bidang.group)
            ws.cell(row=x, column=8, value=invoice.bidang.pemilik_name)
            ws.cell(row=x, column=9, value=invoice.alashak)
            ws.cell(row=x, column=10, value=invoice.bidang.luas_surat or 0).number_format = '0.00'
            ws.cell(row=x, column=11, value=invoice.bidang.luas_bayar or 0).number_format = '0.00'
            ws.cell(row=x, column=12, value=invoice.amount_netto or 0).number_format = '"Rp "#,##0.00_);[Red]("Rp"#,##0.00)'
            ws.cell(row=x, column=13, value=invoice.jenis_bayar)
            ws.cell(row=x, column=14, value=status_workflow)
            ws.cell(row=x, column=15, value=last_status_workflow_at)
            # if (invoice.bidang.is_ptsl or False) == False:
            #     ws.cell(row=x, column=16, value=invoice.bidang.harga_transaksi or 0).number_format = '"Rp "#,##0.00_);[Red]("Rp"#,##0.00)'
            # else:
            #     ws.cell(row=x, column=16, value=invoice.bidang.harga_ptsl or 0).number_format = '"Rp "#,##0.00_);[Red]("Rp"#,##0.00)'
            
            ws.cell(row=x, column=16, value=invoice.bidang.harga_transaksi or 0).number_format = '"Rp "#,##0.00_);[Red]("Rp"#,##0.00)'
                
            ws.cell(row=x, column=17, value=','.join([f'{payment_detail.nomor_giro if payment_detail else {""}} : Rp. {"{:,.0f}".format(payment_detail.amount)}' for payment_detail in invoice.payment_details]))
            ws.cell(row=x, column=18, value=invoice.termin.tanggal_rencana_transaksi)
            ws.cell(row=x, column=19, value=invoice.payment_status if invoice.payment_status else invoice.payment_status_ext)
            ws.cell(row=x, column=20, value=invoice.last_payment_status_at)

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=memo_data.xlsx"})

# SEARCH KJB HD UNTUK UTJ
@router.get("/search/kjb_hd", response_model=GetResponseBaseSch[list[KjbHdSearchSch]])
async def get_list_kjb_hd(
                keyword:str = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(KjbHd)
    query = query.outerjoin(Termin, Termin.kjb_hd_id == KjbHd.id)
    query = query.filter(Termin.kjb_hd_id == None)
    query = query.filter(or_(KjbHd.is_draft != True, KjbHd.is_draft is None))
    
    if keyword:
        query = query.filter(KjbHd.code.ilike(f'%{keyword}%'))


    objs = await crud.kjb_hd.get_multi_no_page(query=query)
    return create_response(data=objs)

# GET SINGLE KJB HD UNTUK UTJ
@router.get("/search/kjb_hd/{id}", response_model=GetResponseBaseSch[KjbHdForTerminByIdSch])
async def get_list_bidang_by_kjb_hd_id(
                id:UUID,
                termin_id:UUID | None = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    obj = await crud.kjb_hd.get_by_id_for_termin(id=id)
    bidang_details = await crud.bidang.get_multi_by_kjb_hd_id(kjb_hd_id=id)

    obj_return = KjbHdForTerminByIdSch(**dict(obj))

    bidangs = []
    for b in bidang_details:
        bidang = BidangForUtjSch(**dict(b))
        bidangs.append(bidang)

    obj_return.bidangs = bidangs

    return create_response(data=obj_return)

# GET LIST TAHAP UNTUK MEMO BAYAR
@router.get("/search/tahap", response_model=GetResponseBaseSch[list[TahapSrcSch]])
async def get_list_tahap(
                keyword: str | None = None,
                jenis_bayar: JenisBayarEnum | None = None, 
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(Tahap)
    query = query.join(TahapDetail, Tahap.id == TahapDetail.tahap_id)
    query = query.join(Spk, TahapDetail.bidang_id == Spk.bidang_id)
    query = query.join(Workflow, and_(Workflow.reference_id == Spk.id, Workflow.entity == WorkflowEntityEnum.SPK))
    query = query.join(Planing, Planing.id == Tahap.planing_id)
    query = query.join(Project, Project.id == Planing.project_id)
    query = query.outerjoin(Invoice, and_(Spk.id == Invoice.spk_id, Invoice.is_void != True))
    query = query.where(and_(Spk.jenis_bayar != JenisBayarEnum.PAJAK,
                            Spk.is_void != True,
                            Invoice.id == None,
                            Workflow.last_status == WorkflowLastStatusEnum.COMPLETED))
    
    if jenis_bayar:
        query = query.filter(Spk.jenis_bayar == jenis_bayar)
    
    if keyword:
        query = query.filter(or_(cast(Tahap.nomor_tahap, String).ilike(f'%{keyword}%'),
                                Project.name.ilike(f'%{keyword}%')))

    query = query.distinct()
    query = query.options(selectinload(Tahap.planing
                        ).options(selectinload(Planing.project))
                )

    objs = await crud.tahap.get_multi_no_page(query=query)
    return create_response(data=objs)

# GET SINGLE TAHAP UNTUK MEMO BAYAR
@router.get("/search/tahap/{id}", response_model=GetResponseBaseSch[list[SpkJenisBayarSch]])
async def get_list_tahap_by_id(
                id: UUID | None = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(Spk).join(TahapDetail, TahapDetail.bidang_id == Spk.bidang_id
                        ).outerjoin(Invoice, and_(Spk.id == Invoice.spk_id, Invoice.is_void != True)
                        ).outerjoin(Workflow, and_(Workflow.reference_id == Spk.id, Workflow.entity == WorkflowEntityEnum.SPK)
                        ).where(and_(TahapDetail.tahap_id == id,
                                    Spk.jenis_bayar != JenisBayarEnum.PAJAK,
                                    Spk.is_void != True,
                                    Invoice.id == None,
                                    Workflow.last_status == WorkflowLastStatusEnum.COMPLETED))

    objs = await crud.spk.get_multi_no_page(query=query)

    spk_jenis_bayars = []
    jenis_bayars = []
    for spk in objs:
        if spk.jenis_bayar not in jenis_bayars:
            spk_ = SpkJenisBayarSch(jenis_bayar=spk.jenis_bayar)
            spk_jenis_bayars.append(spk_)
            jenis_bayars.append(spk.jenis_bayar)

    return create_response(data=spk_jenis_bayars)

@router.get("/search/komponen_biaya", response_model=GetResponseBaseSch[list[BidangKomponenBiayaSch]])
async def get_list_komponen_biaya_by_bidang_id_and_invoice_id(
                bidang_id:UUID,
                invoice_id:UUID,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""
    pengembalian = False
    invoice = await crud.invoice.get_by_id(id=invoice_id)
    if invoice.jenis_bayar == JenisBayarEnum.PENGEMBALIAN_BEBAN_PENJUAL:
        pengembalian = True
    
    biaya_lain = False
    if invoice.jenis_bayar == JenisBayarEnum.BIAYA_LAIN:
        biaya_lain = True
    
    objs = await crud.bidang_komponen_biaya.get_multi_beban_by_invoice_id(invoice_id=invoice_id, pengembalian=pengembalian, biaya_lain=biaya_lain)
    if bidang_id:
        objs_2:list[BidangKomponenBiaya] = []
        datas = await crud.bidang_komponen_biaya.get_multi_beban_by_bidang_id(bidang_id=bidang_id)
        for data in datas:
            if (data.komponen_biaya_outstanding or 0) == 0 and (data.estimated_amount or 0) != 0:
                continue
            else:
                objs_2.append(data)
            
        for obj2 in objs_2:
            obj1 = next((obj for obj in objs if obj.beban_biaya_id == obj2.beban_biaya_id), None)
            if obj1:
                continue
            else:
                objs.append(obj2)

    return create_response(data=objs)

@router.get("/search/komponen_biaya/bidang/{spk_id}", response_model=GetResponseBaseSch[list[BidangKomponenBiayaSch]])
async def get_list_komponen_biaya_by_spk_id(
                spk_id:UUID,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    spk = await crud.spk.get(id=spk_id)

    objs = []
    if spk.jenis_bayar == JenisBayarEnum.PENGEMBALIAN_BEBAN_PENJUAL:
        objs = await crud.bidang_komponen_biaya.get_multi_pengembalian_beban_by_bidang_id(bidang_id=spk.bidang_id)
    elif spk.jenis_bayar == JenisBayarEnum.BIAYA_LAIN:
        objs = await crud.bidang_komponen_biaya.get_multi_beban_biaya_lain_by_bidang_id(bidang_id=spk.bidang_id)
    else:
        datas = await crud.bidang_komponen_biaya.get_multi_beban_by_bidang_id(bidang_id=spk.bidang_id)
        for data in datas:
            if (data.komponen_biaya_outstanding or 0) == 0 and (data.estimated_amount or 0) != 0 and (data.is_retur or False) == False:
                continue
            else:
                objs.append(data)


    return create_response(data=objs)

@router.post("/search/notaris", response_model=GetResponseBaseSch[list[NotarisSrcSch]])
async def get_list_manager(
                sch:TerminBidangIDSch,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    list_id = [id.bidang_id for id in sch.bidangs]

    objs = await crud.notaris.get_multi_by_bidang_ids(list_ids=list_id)

    return create_response(data=objs)

@router.post("/search/manager", response_model=GetResponseBaseSch[list[ManagerSrcSch]])
async def get_list_manager(
                sch:TerminBidangIDSch,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    list_id = [id.bidang_id for id in sch.bidangs]

    objs = await crud.manager.get_multi_by_bidang_ids(list_ids=list_id)

    return create_response(data=objs)

@router.post("/search/sales", response_model=GetResponseBaseSch[list[SalesSrcSch]])
async def get_list_sales(
                sch:TerminBidangIDSch,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    list_id = [id.bidang_id for id in sch.bidangs]

    objs = await crud.sales.get_multi_by_bidang_ids_and_manager_id(list_ids=list_id, manager_id=sch.manager_id)

    return create_response(data=objs)

@router.post("/search/rekening", response_model=GetResponseBaseSch[list[RekeningSch]])
async def get_list_rekening(
                sch:TerminBidangIDSch,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    list_id = [id.bidang_id for id in sch.bidangs]

    objs = await crud.rekening.get_multi_by_bidang_ids(list_ids=list_id)

    return create_response(data=objs)

@router.get("/search/spk", response_model=GetResponseBaseSch[list[SpkSrcSch]])
async def get_list_spk_by_tahap_id(
                tahap_id:UUID | None = None,
                termin_id:UUID | None = None,
                jenis_bayar:JenisBayarEnum | None = None,
                keyword:str = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""
    objs = []
    objs_return = []

    if tahap_id == None and termin_id == None:
        objs = await crud.spk.get_multi_by_keyword_tahap_id_and_termin_id(keyword=keyword)
        objs = [spk for spk in objs if len([invoice for invoice in spk.invoices if invoice.is_void != True]) == 0 and spk.is_void != True]

    if tahap_id and termin_id == None:
        objs_with_tahap = await crud.spk.get_multi_by_keyword_tahap_id_and_termin_id(keyword=keyword, 
                                                                                     tahap_id=tahap_id, 
                                                                                     jenis_bayar=jenis_bayar)
        objs = objs + objs_with_tahap
        # objs = list(set(objs + objs_with_tahap))
    
    if tahap_id and termin_id:
        objs_with_tahap_termin = await crud.spk.get_multi_by_keyword_tahap_id_and_termin_id(keyword=keyword, 
                                                                                            tahap_id=tahap_id, 
                                                                                            termin_id=termin_id,
                                                                                            jenis_bayar=jenis_bayar)
        objs = objs + objs_with_tahap_termin

    
    # objs = [obj for obj in objs if obj.status_workflow == WorkflowLastStatusEnum.COMPLETED]

    reference_ids = [spk.id for spk in objs]
    workflows = await crud.workflow.get_by_reference_ids(reference_ids=reference_ids, entity=WorkflowEntityEnum.SPK)

    for obj in objs:
        workflow = next((wf for wf in workflows if wf.reference_id == obj.id), None)
        last_status = workflow.last_status if workflow else ''
        if last_status == WorkflowLastStatusEnum.COMPLETED:
            objs_return.append(obj)

    return create_response(data=objs_return)

@router.get("/search/spk/{id}", response_model=GetResponseBaseSch[SpkInTerminSch])
async def get_by_id(id:UUID,
                    current_worker:Worker = Depends(crud.worker.get_active_worker)):

    """Get an object by id"""

    spk = await TerminService().get_spk_by_id(spk_id=id)

    return create_response(data=spk)

@router.post("/search/spk/by-ids", response_model=GetResponseBaseSch[list[SpkInTerminSch]])
async def get_by_ids(sch:SpkIdSch,
                    current_worker:Worker = Depends(crud.worker.get_active_worker)):

    """Get an object by ids"""

    spks = []
    for id in sch.spk_ids:
        spk = await TerminService().get_spk_by_id(spk_id=id)
        spks.append(spk)

    return create_response(data=spks)
    
@router.get("/history/memo/{bidang_id}", response_model=GetResponseBaseSch[list[InvoiceHistoryInTermin]])
async def get_list_history_memo_by_bidang_id(bidang_id:UUID,
                                            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Get list history memo bayar dp by bidang_id"""

    objs = await crud.invoice.get_multi_history_invoice_by_bidang_id(bidang_id=bidang_id)

    return create_response(data=objs)


# TASK WORKFLOW
@router.post("/task-workflow")
async def create_workflow(payload:Dict, request:Request):
    
    id = payload.get("id", None)
    obj = await crud.termin.get(id=id)

    if not obj:
        raise IdNotFoundException(Spk, id)
    
    await TerminService().task_workflow(obj=obj, request=request)

    return create_response({"detail": "SUCCESS"})

# TASK CREATE RFP
@router.post("/task/create_rfp")
async def create_rfp(payload: Dict):

    data, msg = await RfpService().create_rfp(termin_id=payload["id"])

    if data is not None:
        termin_current = await crud.termin.get(id=data.client_ref_no)
        if termin_current is None:
            raise HTTPException(status_code=404, detail="Memo Bayar not found!")
        
        termin_updated = TerminUpdateBaseSch.from_orm(termin_current)
        termin_updated.rfp_ref_no = data.id
        termin_updated.rfp_last_status = data.current_step

        await crud.termin.update(obj_current=termin_current, obj_new=termin_updated)
    else:
        raise HTTPException(status_code=409, detail=msg)

    return create_response({"detail": "SUCCESS"})