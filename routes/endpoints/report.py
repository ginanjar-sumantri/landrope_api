from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlmodel import text, select, SQLModel, update
from fastapi_pagination import Params, Page
from fastapi_pagination.ext.async_sqlalchemy import paginate
from fastapi_async_sqlalchemy import db
from models import HasilPetaLokasi
from schemas.hasil_peta_lokasi_sch import HasiLPetaLokasiUpdateTanggalKirimBerkasSch
from schemas.report_sch import KekuranganBerkasManagerSch
from schemas.response_sch import (GetResponseBaseSch, GetResponsePaginatedSch, create_response)
from common.rounder import RoundTwo
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.writer.excel import save_workbook
from datetime import date, datetime
from io import BytesIO
import crud

router = APIRouter()

@router.get("/tim-ukur")
async def report_tim_ukur(start_date:date | None = None, end_date:date|None = None):

    wb = Workbook()
    ws = wb.active

    ws.title =  "REPORT TIM UKUR"
    ws.firstHeader

    ws.cell(row=1, column=2, value="LAPORAN PENGUKURAN")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.cell(row=2, column=2, value=f"CUT OFF DATE {start_date} S/D {end_date}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)

    header_string = ["No", "No. Permintaan Ukur", "Tanggal Terima Berkas", "Mediator", "Group", "Desa", "Nama Pemilik Tanah", "Jenis Surat (GIRIK/SHM)", 
                    "Alas Hak", "Luas Surat (m2)", "Tanggal Pengukuran", "Penunjuk Batas", "Luas Ukur (m2)", "Surveyor", "Tanggal Kirim Ukur ke Analis", "Keterangan"]
    
    end_column = len(header_string)

    for idx, val in enumerate(header_string):
        ws.cell(row=3, column=idx + 1, value=val).font = Font(bold=True)
        if idx + 1 < 11:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
        else:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    query_start_date = ""
    if start_date and end_date:
        query_start_date = "WHERE rpl.tanggal_terima_berkas >=" f"'{start_date}'"
        query_start_date += " AND rpl.tanggal_terima_berkas <" f"'{end_date}'"

    query = f"""
            SELECT
            ROW_NUMBER() OVER (ORDER BY rpl.code) AS no,
            rpl.code as no_permintaan_ukur,
            rpl.tanggal_terima_berkas,
            hd.mediator,
            dt.group,
            d.name as desa_name,
            p.name as pemilik_name,
            dt.jenis_alashak,
            dt.alashak,
            dt.luas_surat as luas_surat,
            rpl.tanggal_pengukuran,
            rpl.penunjuk_batas,
            rpl.luas_ukur,
            rpl.surveyor,
            rpl.tanggal_kirim_ukur,
            ket.name as keterangan
            FROM request_peta_lokasi rpl
            INNER JOIN kjb_dt dt ON dt.id = rpl.kjb_dt_id
            INNER JOIN kjb_hd hd ON hd.id = dt.kjb_hd_id
            LEFT OUTER JOIN desa d ON d.id = dt.desa_by_ttn_id
            LEFT OUTER JOIN pemilik p ON p.id = dt.pemilik_id
            LEFT OUTER JOIN keterangan_req_petlok ket ON ket.id = rpl.keterangan_req_petlok_id
            {query_start_date}
    """

    db_session = db.session
    response = await db_session.execute(query)
    result = response.all()

    x = 3
    for row_data in result:
        x += 1
        ws.cell(row=x, column=1, value=row_data[0])
        ws.cell(row=x, column=2, value=row_data[1])
        ws.cell(row=x, column=3, value=row_data[2])
        ws.cell(row=x, column=4, value=row_data[3])
        ws.cell(row=x, column=5, value=row_data[4])
        ws.cell(row=x, column=6, value=row_data[5])
        ws.cell(row=x, column=7, value=row_data[6])
        ws.cell(row=x, column=8, value=row_data[7])
        ws.cell(row=x, column=9, value=row_data[8])
        ws.cell(row=x, column=10, value=row_data[9])
        ws.cell(row=x, column=11, value=row_data[10])
        ws.cell(row=x, column=12, value=row_data[11])
        ws.cell(row=x, column=13, value=row_data[12])
        ws.cell(row=x, column=14, value=row_data[13])
        ws.cell(row=x, column=15, value=row_data[14])
        ws.cell(row=x, column=16, value=row_data[15])
    
    ws.cell(row=x + 2, column=2, value="CATATAN:")
    ws.merge_cells(start_row=x + 2, start_column=2, end_row=x + 2, end_column=end_column)
    ws.cell(row=x + 3, column=2, value="1. KOLOM B S/D J DIPEROLEH DARI TIM MARKETING (DARI REQUEST UKUR/PETA LOKASI)")
    ws.merge_cells(start_row=x + 3, start_column=2, end_row=x + 3, end_column=end_column)
    ws.cell(row=x + 4, column=2, value="2. KOLOM K S/D P DIISI OLEH ADMIN UKUR DI SISTEM, SETELAH MENERIMA HASIL UKUR")
    ws.merge_cells(start_row=x + 4, start_column=2, end_row=x + 4, end_column=end_column)
    ws.cell(row=x + 5, column=2, value="3. 'TANGGAL PENGUKURAN' ADALAH TANGGAL TIM UKUR MELAKUKAN PENGUKURAN DI LAPANGAN ")
    ws.merge_cells(start_row=x + 5, start_column=2, end_row=x + 5, end_column=end_column)
    ws.cell(row=x + 6, column=2, value="4. 'TANGGAL KIRIM UKUR KE ANALIS' ADALAH TANGGAL TIM UKUR MENGIRIMKAN HASIL UKUR KE ANALIS")
    ws.merge_cells(start_row=x + 6, start_column=2, end_row=x + 6, end_column=end_column)
    ws.cell(row=x + 7, column=2, value="5. 'KETERANGAN' DIISI DENGAN INFORMASI DARI TIM UKUR APABILA ADA KENDALA/PENDING SEPERTI : BELUM UKUR MENUNGGU INFO MEDIATOR, PENJUAL BELUM MENUNJUKKAN BIDANG, DLL")
    ws.merge_cells(start_row=x + 7, start_column=2, end_row=x + 7, end_column=end_column)
    ws.cell(row=x + 8, column=2, value="6. 'TANGGAL TERIMA BERKAS' ADALAH TANGGAL TIM UKUR MENERIMA BERKAS UKUR DARI MARKETING")
    ws.merge_cells(start_row=x + 8, start_column=2, end_row=x + 8, end_column=end_column)
    

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=tim_ukur.xlsx"})

@router.get("/tim-analyst")
async def report_tim_analyst(start_date:date | None = None, end_date:date|None = None):

    wb = Workbook()
    ws = wb.active

    ws.title =  "REPORT TIM ANALYST"
    ws.firstHeader

    ws.cell(row=1, column=2, value="LAPORAN ANALISA PETA LOKASI")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.cell(row=2, column=2, value=f"CUT OFF DATE {start_date} S/D {end_date}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)

    header_string = ["No", "No. Permintaan Ukur", "Tanggal Serah Terima Ukur", "Mediator", "Group", "Nama Pemilik Tanah", "Jenis Surat", "Alas Hak", "Luas (m2)", 
                    "Desa", "Project", "PT SK", "Nama Petugas Analyst", "No. ID Bidang", "L SURAT", "L UKUR", "L NETT", "L CLEAR", "L OVERLAP", "ID OVERLAP", 
                    "Ket (Overlap/Clear)", "L Proses (PBT dan SERTIFIKASI)", "Tanggal Serah Terima ke Tim Marketing", "Selisih Hari"]
    
    end_column = len(header_string)

    for idx, val in enumerate(header_string):
        ws.cell(row=3, column=idx + 1, value=val).font = Font(bold=True)
        if idx + 1 < 10:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        else: 
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
            

    query_start_date = ""
    if start_date and end_date:
        query_start_date = "WHERE rpl.tanggal_terima_berkas >=" f"'{start_date}'"
        query_start_date += " AND rpl.tanggal_terima_berkas <" f"'{end_date}'"

    query = f"""
            SELECT
                ROW_NUMBER() OVER (ORDER BY rpl.code) AS no,
                rpl.code as no_permintaan_ukur,
                rpl.tanggal_kirim_ukur,
                hd.mediator,
                dt.group,
                p.name as pemilik_name,
                dt.jenis_alashak,
                dt.alashak,
                dt.luas_surat as luas_surat,
                d.name as desa_name,
                pr.name as project_name,
                pt.name as ptsk_name,
                w.name as petugas_name,
                b.id_bidang as no_id_bidang,
                hpl.luas_surat,
                hpl.luas_ukur,
                hpl.luas_nett,
                hpl.luas_clear,
                bo.luas as luas_overlap,
                bd.id_bidang as id_overlap,
                hpl.hasil_analisa_peta_lokasi,
                hpl.luas_proses,
                hpl.tanggal_kirim_berkas,
                Coalesce((hpl.tanggal_kirim_berkas - rpl.tanggal_terima_berkas), 0) as selisih
            FROM 
                request_peta_lokasi rpl 
            INNER JOIN 
                hasil_peta_lokasi hpl ON rpl.id = hpl.request_peta_lokasi_id 
            LEFT OUTER JOIN 
                hasil_peta_lokasi_detail hpl_dt ON hpl.id = hpl_dt.hasil_peta_lokasi_id
            LEFT OUTER JOIN 
                bidang_overlap bo ON bo.id = hpl_dt.bidang_overlap_id
            INNER JOIN 
                kjb_dt dt ON dt.id = rpl.kjb_dt_id
            INNER JOIN 
                kjb_hd hd ON hd.id = dt.kjb_hd_id
            LEFT OUTER JOIN 
                pemilik p ON p.id = dt.pemilik_id
            LEFT OUTER JOIN 
                planing pl ON pl.id = hpl.planing_id
            LEFT OUTER JOIN 
                desa d ON d.id = pl.desa_id
            LEFT OUTER JOIN 
                project pr ON pr.id = pl.project_id
            LEFT OUTER JOIN 
                skpt sk ON sk.id = hpl.skpt_id
            LEFT OUTER JOIN 
                ptsk pt ON pt.id = sk.ptsk_id
            INNER JOIN 
                worker w ON w.id = hpl.created_by_id
            INNER JOIN 
                bidang b ON b.id = hpl.bidang_id
            LEFT OUTER JOIN 
                bidang bd ON bd.id = bo.parent_bidang_intersect_id
            {query_start_date}
    """

    db_session = db.session
    response = await db_session.execute(query)
    result = response.all()

    x = 3
    for row_data in result:
        x += 1
        ws.cell(row=x, column=1, value=row_data[0])
        ws.cell(row=x, column=2, value=row_data[1])
        ws.cell(row=x, column=3, value=row_data[2])
        ws.cell(row=x, column=4, value=row_data[3])
        ws.cell(row=x, column=5, value=row_data[4])
        ws.cell(row=x, column=6, value=row_data[5])
        ws.cell(row=x, column=7, value=row_data[6])
        ws.cell(row=x, column=8, value=row_data[7])
        ws.cell(row=x, column=9, value=row_data[8])
        ws.cell(row=x, column=10, value=row_data[9])
        ws.cell(row=x, column=11, value=row_data[10])
        ws.cell(row=x, column=12, value=row_data[11])
        ws.cell(row=x, column=13, value=row_data[12])
        ws.cell(row=x, column=14, value=row_data[13])
        ws.cell(row=x, column=15, value=row_data[14])
        ws.cell(row=x, column=16, value=row_data[15])
        ws.cell(row=x, column=17, value=row_data[16])
        ws.cell(row=x, column=18, value=row_data[17])
        ws.cell(row=x, column=19, value=row_data[18])
        ws.cell(row=x, column=20, value=row_data[19])
        ws.cell(row=x, column=21, value=row_data[20])
        ws.cell(row=x, column=22, value=row_data[21])
        ws.cell(row=x, column=23, value=row_data[22])
        ws.cell(row=x, column=24, value=row_data[23])
    
    ws.cell(row=x + 2, column=2, value="CATATAN:")
    ws.merge_cells(start_row=x + 2, start_column=2, end_row=x + 2, end_column=end_column)
    ws.cell(row=x + 3, column=2, value="LUAS NETT: LUAS UKUR YANG DIKURANGI DENGAN OVERLAP BID LAIN NON BINTANG SEPERTI: OVERLAP SHM, DAN OVERLAP TANAH MILIK PT")
    ws.merge_cells(start_row=x + 3, start_column=2, end_row=x + 3, end_column=end_column)
    ws.cell(row=x + 4, column=2, value="LUAS SURAT ADALAH LUAS YANG DIDAPAT DARI ALAS HAK YANG DITERIMA. SUDAH DIINPUT OLEH TIM MARKETING DARI MODUL PRA-PEMBEBASAN")
    ws.merge_cells(start_row=x + 4, start_column=2, end_row=x + 4, end_column=end_column)
    ws.cell(row=x + 5, column=2, value="LUAS UKUR ADALAH LUAS HASIL UKUR FISIK LAPANGAN")
    ws.merge_cells(start_row=x + 5, start_column=2, end_row=x + 5, end_column=end_column)
    ws.cell(row=x + 6, column=2, value="LUAS SURAT BISA BERUBAH, KASUSNYA LUAS FISIK JAUH LEBIH BESAR DARI LUAS SURAT SEHINGGA")
    ws.merge_cells(start_row=x + 6, start_column=2, end_row=x + 6, end_column=end_column)
    ws.cell(row=x + 7, column=2, value="ADA MODUL UNTUK REVISI LUAS SURAT DI TIM PRA PEMBEBASAN, ADA PILIHAN UNTUK RENVOY SURAT")
    ws.merge_cells(start_row=x + 7, start_column=2, end_row=x + 7, end_column=end_column)

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=tim_analyst.xlsx"})

@router.post("/hasil-analisa")
async def report_hasil_analisa(background_task:BackgroundTasks, sch:HasiLPetaLokasiUpdateTanggalKirimBerkasSch|None = None):
    
    ids = [f"'{str(req)}'" for req in sch.ids]
    ids_str = ",".join(ids)
    
    wb = Workbook()
    ws = wb.active

    ws.title =  "REPORT HASIL ANALISA"
    ws.firstHeader

    ws.cell(row=1, column=2, value="HASIL ANALISA")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.cell(row=2, column=2, value=f"NO: {sch.code}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)

    header_string = ["No", "Mediator", "Group", "Nama Pemilik Tanah", "No. Telp Pemilik Tanah", "Alas Hak", "Luas (m2)", 
                    "Desa", "Project", "PT", "Nama Petugas Analyst", "No. Id Bidang", "Luas (m2)", "Ket (Overlap/Clear)"]

    for idx, val in enumerate(header_string):
        ws.cell(row=4, column=idx + 1, value=val).font = Font(bold=True)
        ws.cell(row=4, column=idx + 1, value=val).fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")


    query = f"""
            SELECT 
            ROW_NUMBER() OVER (ORDER BY hpl.id) AS no,
            hd.mediator,
            dt.group,
            p.name as pemilik_name,
            (select nomor_telepon from kontak k where k.pemilik_id = p.id limit 1) as no_pemilik,
            dt.alashak,
            dt.luas_surat_by_ttn as luas_surat,
            d.name as desa_name,
            pr.name as project_name,
            pt.name as ptsk_name,
            w.name as petugas_name,
            b.id_bidang,
            hpl.luas_ukur,
            hpl.hasil_analisa_peta_lokasi
            FROM hasil_peta_lokasi hpl
            INNER JOIN kjb_dt dt ON dt.id = hpl.kjb_dt_id 
            INNER JOIN kjb_hd hd ON hd.id = dt.kjb_hd_id
            INNER JOIN pemilik p ON p.id = hpl.pemilik_id
            LEFT OUTER JOIN planing pl ON pl.id = hpl.planing_id
            LEFT OUTER JOIN desa d ON d.id = pl.desa_id
            LEFT OUTER JOIN project pr ON pr.id = pl.project_id
            LEFT OUTER JOIN skpt sk ON sk.id = hpl.skpt_id
            LEFT OUTER JOIN ptsk pt ON pt.id = sk.ptsk_id
            INNER JOIN worker w ON w.id = hpl.created_by_id
            INNER JOIN bidang b ON b.id = hpl.bidang_id
            WHERE hpl.id IN ({ids_str})
    """

    db_session = db.session
    response = await db_session.execute(query)
    result = response.all()

    x = 4
    for row_data in result:
        x += 1
        ws.cell(row=x, column=1, value=row_data[0])
        ws.cell(row=x, column=2, value=row_data[1])
        ws.cell(row=x, column=3, value=row_data[2])
        ws.cell(row=x, column=4, value=row_data[3])
        ws.cell(row=x, column=5, value=row_data[4])
        ws.cell(row=x, column=6, value=row_data[5])
        ws.cell(row=x, column=7, value=row_data[6])
        ws.cell(row=x, column=8, value=row_data[7])
        ws.cell(row=x, column=9, value=row_data[8])
        ws.cell(row=x, column=10, value=row_data[9])
        ws.cell(row=x, column=11, value=row_data[10])
        ws.cell(row=x, column=12, value=row_data[11])
        ws.cell(row=x, column=13, value=row_data[12])
        ws.cell(row=x, column=14, value=row_data[13])

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)

    background_task.add_task(update_tanggal_kirim_berkas, sch)

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=hasil_analisa.xlsx"})

async def update_tanggal_kirim_berkas(sch:HasiLPetaLokasiUpdateTanggalKirimBerkasSch):
     
    query = update(HasilPetaLokasi).where(HasilPetaLokasi.id.in_(sch.ids)).values(tanggal_kirim_berkas=sch.tanggal_kirim_berkas)
    
    db_session = db.session
    await db_session.execute(query)
    await db_session.commit()

@router.get("/summary-analyst")
async def report_summary_analyst(start_date:date|None = None, end_date:date|None = None):

    wb = Workbook()
    ws = wb.active

    ws.title =  "REPORT TIM UKUR"
    ws.firstHeader

    ws.cell(row=1, column=2, value="RESUME LAPORAN TIM ANALISA DAN TIM AKUR")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    ws.cell(row=2, column=2, value=f"CUT OFF DATE {start_date} S/D {end_date}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)

    ws.cell(row=3, column=8, value="Laporan Ukuran")
    ws.cell(row=3, column=8, value="Laporan Ukuran").font = Font(bold=True)
    ws.cell(row=3, column=8, value="Laporan Ukuran").fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws.cell(row=3, column=8, value="Laporan Ukuran").alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=11)

    ws.cell(row=3, column=12, value=f"Laporan Analis")
    ws.cell(row=3, column=12, value="Laporan Analis").font = Font(bold=True)
    ws.cell(row=3, column=12, value="Laporan Analis").fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
    ws.cell(row=3, column=12, value="Laporan Analis").alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=12, end_row=3, end_column=15)

    r4c8 = ws.cell(row=4, column=8)
    r4c8.font = Font(bold=True)
    r4c8.fill = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type="solid")
    r4c8.alignment = Alignment(horizontal='center', vertical='center')
    r4c8.value = "Sudah Ukur"
    ws.merge_cells(start_row=4, start_column=8, end_row=4, end_column=9)

    r4c10 = ws.cell(row=4, column=10)
    r4c10.font = Font(bold=True)
    r4c10.fill = PatternFill(start_color="00993366", end_color="00993366", fill_type="solid")
    r4c10.alignment = Alignment(horizontal='center', vertical='center')
    r4c10.value = "Belum Ukur"
    ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=11)

    r4c12 = ws.cell(row=4, column=12)
    r4c12.font = Font(bold=True)
    r4c12.fill = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type="solid")
    r4c12.alignment = Alignment(horizontal='center', vertical='center')
    r4c12.value = "Dibeli"
    ws.merge_cells(start_row=4, start_column=12, end_row=4, end_column=13)

    r4c14 = ws.cell(row=4, column=14)
    r4c14.font = Font(bold=True)
    r4c14.fill = PatternFill(start_color="00993366", end_color="00993366", fill_type="solid")
    r4c14.alignment = Alignment(horizontal='center', vertical='center')
    r4c14.value = "Tidak Dibeli"
    ws.merge_cells(start_row=4, start_column=14, end_row=4, end_column=15)

    header_string = ["Nomor KJB", "Desa", "Group", "Manager", "No Permintaan Ukur", "Jumlah Bidang", "Total Luas (m2)",
                    "Bidang", "Luas (m2)", "Bidang", "Luas (m2)", "Bidang", "Luas (m2)", "Bidang", "Luas (m2)", "Status"]
    
    end_column = len(header_string)

    for idx, val in enumerate(header_string):
        ws.cell(row=5, column=idx + 1, value=val).font = Font(bold=True)
        ws.cell(row=5, column=idx + 1, value=val).fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")

    query_start_date = ""
    if start_date and end_date:
        query_start_date = "WHERE rpl.tanggal_terima_berkas >=" f"'{start_date}'"
        query_start_date += " AND rpl.tanggal_terima_berkas <" f"'{end_date}'"

    query = f"""
            WITH subquery AS (SELECT
                hd.code AS nomor_kjb,
                d.name AS desa_name,
                dt.group,
                m.name AS manager_name,
                rpl.code AS nomor_request,
                COUNT(rpl.kjb_dt_id) AS jumlah_bidang,
                SUM(dt.luas_surat_by_ttn) AS total_luas,
                COALESCE((SELECT COUNT(rpl_.id) 
                FROM request_peta_lokasi rpl_ 
                INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
                WHERE rpl_.code = rpl.code 
                AND d.id = dt_.desa_by_ttn_id 
                AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
                AND rpl_.luas_ukur IS NOT NULL 
                AND rpl_.tanggal_pengukuran IS NOT NULL), 0) AS sudah_diukur,
            COALESCE((SELECT SUM(rpl_.luas_ukur) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
            AND rpl_.luas_ukur IS NOT NULL 
            AND rpl_.tanggal_pengukuran IS NOT NULL), 0) AS luas_sudah_diukur,
            COALESCE((SELECT COUNT(rpl_.id) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            WHERE rpl_.code = rpl.code
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
            AND rpl_.luas_ukur IS NULL 
            AND rpl_.tanggal_pengukuran IS NULL), 0) AS belum_diukur,
            COALESCE((SELECT SUM(dt_.luas_surat_by_ttn) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
            AND rpl_.luas_ukur IS NULL 
            AND rpl_.tanggal_pengukuran IS NULL), 0) AS luas_belum_diukur,
            COALESCE((SELECT COUNT(rpl_.id) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            INNER JOIN hasil_peta_lokasi hpl_ ON rpl_.id = hpl_.request_peta_lokasi_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')), 0) AS dibeli,
            COALESCE((SELECT SUM(hpl_.luas_surat) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            INNER JOIN hasil_peta_lokasi hpl_ ON rpl_.id = hpl_.request_peta_lokasi_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')), 0) AS luas_dibeli,
            COALESCE((SELECT COUNT(rpl_.id) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            LEFT OUTER JOIN hasil_peta_lokasi hpl_ ON rpl_.id = hpl_.request_peta_lokasi_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
            AND hpl_.id IS NULL), 0) AS tidak_dibeli,
            COALESCE((SELECT SUM(hpl_.luas_surat) 
            FROM request_peta_lokasi rpl_ 
            INNER JOIN kjb_dt dt_ ON dt_.id = rpl_.kjb_dt_id
            INNER JOIN hasil_peta_lokasi hpl_ ON rpl_.id = hpl_.request_peta_lokasi_id
            WHERE rpl_.code = rpl.code 
            AND d.id = dt_.desa_by_ttn_id 
            AND COALESCE(dt.group, '') = COALESCE(dt_.group, '')
            AND hpl_.id IS NULL), 0) AS luas_tidak_dibeli
            FROM request_peta_lokasi rpl
            INNER JOIN kjb_dt dt ON dt.id = rpl.kjb_dt_id
            INNER JOIN kjb_hd hd ON hd.id = dt.kjb_hd_id
            LEFT OUTER JOIN desa d ON d.id = dt.desa_by_ttn_id
            LEFT OUTER JOIN manager m ON m.id = hd.manager_id
            {query_start_date}
            GROUP BY hd.code, d.id, dt.group, m.name, rpl.code)
            SELECT s.*,
            CASE
                WHEN sudah_diukur >= jumlah_bidang THEN 'ORDER UKUR SELESAI'
                ELSE 'ORDER UKUR BELUM SELESAI'
            END AS status
            FROM subquery s
            ORDER BY nomor_request ASC
    """

    db_session = db.session
    response = await db_session.execute(query)
    result = response.all()

    x = 5
    for row_data in result:
        x += 1
        ws.cell(row=x, column=1, value=row_data[0])
        ws.cell(row=x, column=2, value=row_data[1])
        ws.cell(row=x, column=3, value=row_data[2])
        ws.cell(row=x, column=4, value=row_data[3])
        ws.cell(row=x, column=5, value=row_data[4])
        ws.cell(row=x, column=6, value=row_data[5])
        ws.cell(row=x, column=7, value=row_data[6])
        ws.cell(row=x, column=8, value=row_data[7])
        ws.cell(row=x, column=9, value=row_data[8])
        ws.cell(row=x, column=10, value=row_data[9])
        ws.cell(row=x, column=11, value=row_data[10])
        ws.cell(row=x, column=12, value=row_data[11])
        ws.cell(row=x, column=13, value=row_data[12])
        ws.cell(row=x, column=14, value=row_data[13])
        ws.cell(row=x, column=15, value=row_data[14])
        ws.cell(row=x, column=16, value=row_data[15])
    
    total_order_selesai = RoundTwo(sum([data[6] for data in result if data[15] == "ORDER UKUR SELESAI"])/10000)
    total_order_belum_selesai = RoundTwo(sum([data[6] for data in result if data[15] == "ORDER UKUR BELUM SELESAI"])/10000)

    ws.cell(row=x + 3, column=2, value=f"TOTAL ORDER SELESAI: {total_order_selesai} Ha")
    ws.merge_cells(start_row=x + 3, start_column=2, end_row=x + 3, end_column=end_column)
    ws.cell(row=x + 4, column=2, value=f"TOTAL ORDER BELUM SELESAI: {total_order_belum_selesai} Ha")
    ws.merge_cells(start_row=x + 4, start_column=2, end_row=x + 4, end_column=end_column)

    
    ws.cell(row=x + 6, column=2, value="CATATAN:")
    ws.merge_cells(start_row=x + 6, start_column=2, end_row=x + 6, end_column=end_column)
    ws.cell(row=x + 7, column=2, value="1. ORDER UKUR SELESAI JIKA SUDAH JADI PETA LOKASI")
    ws.merge_cells(start_row=x + 7, start_column=2, end_row=x + 7, end_column=end_column)
    ws.cell(row=x + 8, column=2, value="2. JIKA BIDANG SUDAH DIUKUR, NAMUN BELUM JADI PETA LOKASI MAKA MASIH MENJADI OUTSTANDING ORDER UKUR")
    ws.merge_cells(start_row=x + 8, start_column=2, end_row=x + 8, end_column=end_column)
    
    

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=summary_analyst.xlsx"})

######################################################################

@router.get("/kekurangan-berkas-manager")
async def report_kekurangan_berkas_per_manager():

    wb = Workbook()
    ws = wb.active

    ws.title =  "RINCIAN BIDANG KURANG IDENTITAS"
    ws.firstHeader

    ws.cell(row=1, column=2, value="RINCIAN BIDANG KURANG IDENTITAS").font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)

    r2c7 = ws.cell(row=2, column=7)
    r2c7.font = Font(bold=True)
    r2c7.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    r2c7.alignment = Alignment(horizontal='center', vertical='center')
    r2c7.value = "PERORANGAN"
    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=10)

    r2c11 = ws.cell(row=2, column=11)
    r2c11.font = Font(bold=True)
    r2c11.fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
    r2c11.alignment = Alignment(horizontal='center', vertical='center')
    r2c11.value = "PT"
    ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=13)

    r2c14 = ws.cell(row=2, column=14)
    r2c14.font = Font(bold=True)
    r2c14.fill = PatternFill(start_color="00008000", end_color="00008000", fill_type="solid")
    r2c14.alignment = Alignment(horizontal='center', vertical='center')
    r2c14.value = "WARIS"
    ws.merge_cells(start_row=2, start_column=14, end_row=2, end_column=18)

    r2c19 = ws.cell(row=2, column=19)
    r2c19.font = Font(bold=True)
    r2c19.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    r2c19.alignment = Alignment(horizontal='center', vertical='center')
    r2c19.value = "TGL FOLLOW UP"
    ws.merge_cells(start_row=2, start_column=19, end_row=2, end_column=20)

    header_string = ["NO", "DESA", "NAMA PENJUAL", "ALAS HAK", "LUAS (M2)", "LUAS (HA)", 
                    "KTP PENJUAL", "KTP PASANGAN (SUAMI/ISTRI)", "KK", "SURAT NIKAH",
                    "KTP PENGURUS", "AD/ART", "PERUBAHAN AD/ART",
                    "SURAT/AKTA KEMATIAN", "SURAT KETERANGAN AHLI WARIS", "SURAT KUASA WARIS", "KTP SELURUH WARIS", "KK SELURUH WARIS",
                    "SALES/MEDIATOR", "NOTARIS", "KETERANGAN"]
    
    end_column = len(header_string)

    for idx, val in enumerate(header_string):
        ws.cell(row=3, column=idx + 1, value=val).font = Font(bold=True)
        if idx + 1 < 11:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
        elif idx + 1 < 14:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")
        elif idx + 1 < 19:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00008000", end_color="00008000", fill_type="solid")
        else:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
        
        ws.cell(row=3, column=idx + 1, value=val).alignment = Alignment(horizontal='center', vertical='center')

    query = f"""
            WITH perorangan_check AS (
                SELECT 
                    b_hd.id AS bundle_hd_id,
                    dk.name,
                    CASE 
                        WHEN b_dt.meta_data IS NULL THEN 'V' 
                        ELSE ''
                    END AS meta_data
                FROM 
                    bundle_hd b_hd
                CROSS JOIN 
                    (VALUES ('KTP PENJUAL'), ('KTP PASANGAN'), ('KARTU KELUARGA'), ('AKTA NIKAH/SURAT KETERANGAN NIKAH/SURAT CERAI')) AS dkm(name)
                LEFT JOIN 
                    bundle_dt b_dt ON b_dt.bundle_hd_id = b_hd.id
                LEFT JOIN 
                    dokumen dk ON dk.id = b_dt.dokumen_id AND dk.name = dkm.name
                WHERE dk.name IS NOT NULL
            ),
            pt_check AS (
                SELECT 
                    b_hd.id AS bundle_hd_id,
                    dk.name,
                    CASE 
                        WHEN b_dt.meta_data IS NULL THEN 'V' 
                        ELSE '' 
                    END AS meta_data
                FROM 
                    bundle_hd b_hd
                CROSS JOIN 
                    (VALUES ('KTP PENGURUS'), ('AD/ART'), ('PERUBAHAN AD/ART')) AS dkm(name)
                LEFT JOIN 
                    bundle_dt b_dt ON b_dt.bundle_hd_id = b_hd.id
                LEFT JOIN 
                    dokumen dk ON dk.id = b_dt.dokumen_id AND dk.name = dkm.name
                WHERE dk.name IS NOT NULL
            ),
            waris_check AS (
                SELECT 
                    b_hd.id AS bundle_hd_id,
                    dk.name,
                    CASE WHEN dk.name IN ('AKTA KEMATIAN', 'SURAT KEMATIAN', 'SURAT KETERANGAN AHLI WARIS', 'SURAT KUASA AHLI WARIS') THEN
                            CASE 
                                WHEN b_dt.meta_data IS NULL THEN 'V' 
                                ELSE '' 
                            END 
                        WHEN dk.name IN ('KTP AHLI WARIS', 'KK AHLI WARIS') THEN
                            CASE 
                                WHEN COALESCE(dt.jumlah_waris, 0) > COALESCE(b_dt.multiple_count, 0) THEN 'V' 
                                ELSE '' 
                            END 
                        ELSE ''
                    END AS meta_data
                FROM 
                    bundle_hd b_hd
                CROSS JOIN 
                    (VALUES ('AKTA KEMATIAN'), ('SURAT KEMATIAN'), ('SURAT KETERANGAN AHLI WARIS'), ('SURAT KUASA AHLI WARIS'), ('KTP AHLI WARIS'), ('KK AHLI WARIS')) AS dkm(name)
                LEFT JOIN 
                    bundle_dt b_dt ON b_dt.bundle_hd_id = b_hd.id
                LEFT JOIN 
                    dokumen dk ON dk.id = b_dt.dokumen_id AND dk.name = dkm.name
                INNER JOIN
                    kjb_dt dt ON dt.bundle_hd_id = b_hd.id
                WHERE dk.name IS NOT NULL
            ),
            main_query AS (
                SELECT 
                    m.id AS manager_id,
                    m.name AS manager_name,
                    d.name AS desa_name,
                    p.name AS nama_penjual,
                    dt.alashak,
                    COALESCE(dt.luas_surat_by_ttn, 0) AS luas_m2,
                    ROUND(COALESCE(dt.luas_surat_by_ttn / 10000, 0), 2) AS luas_ha,
                    CASE WHEN hd.kategori_penjual = 'Perorangan' THEN dk_ktp.meta_data ELSE '' END AS ktp_penjual,
                    CASE WHEN hd.kategori_penjual = 'Perorangan' THEN dk_pasangan.meta_data ELSE '' END AS ktp_pasangan,
                    CASE WHEN hd.kategori_penjual = 'Perorangan' THEN dk_kk.meta_data ELSE '' END AS kk,
                    CASE WHEN hd.kategori_penjual = 'Perorangan' THEN dk_surat_nikah.meta_data ELSE '' END AS surat_nikah,
                    CASE WHEN hd.kategori_penjual = 'PT' THEN pt_ktp.meta_data ELSE '' END AS ktp_pengurus,
                    CASE WHEN hd.kategori_penjual = 'PT' THEN pt_ad_art.meta_data ELSE '' END AS ad_art,
                    CASE WHEN hd.kategori_penjual = 'PT' THEN pt_p_ad_art.meta_data ELSE '' END AS perubahan_ad_art,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_s_kematian.meta_data ELSE '' END AS surat_kematian,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_a_kematian.meta_data ELSE '' END AS akta_kematian,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_ket_ahli.meta_data ELSE '' END AS surat_ket_ahli_waris,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_kuasa.meta_data ELSE '' END AS surat_kuasa_waris,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_ktp.meta_data ELSE '' END AS ktp_seluruh_waris,
                    CASE WHEN hd.kategori_penjual = 'Waris' THEN waris_kk.meta_data ELSE '' END AS kk_seluruh_waris
                FROM 
                    kjb_dt dt
                INNER JOIN 
                    bundle_hd b_hd ON b_hd.id = dt.bundle_hd_id
                INNER JOIN 
                    kjb_hd hd ON hd.id = dt.kjb_hd_id
                LEFT JOIN 
                    desa d ON d.id = dt.desa_by_ttn_id
                INNER JOIN
                    manager m ON m.id = hd.manager_id
                LEFT JOIN 
                    pemilik p ON p.id = dt.pemilik_id
                LEFT JOIN 
                    perorangan_check dk_ktp ON dk_ktp.bundle_hd_id = b_hd.id AND dk_ktp.name = 'KTP PENJUAL'
                LEFT JOIN 
                    perorangan_check dk_pasangan ON dk_pasangan.bundle_hd_id = b_hd.id AND dk_pasangan.name = 'KTP PASANGAN'
                LEFT JOIN 
                    perorangan_check dk_kk ON dk_kk.bundle_hd_id = b_hd.id AND dk_kk.name = 'KARTU KELUARGA'
                LEFT JOIN 
                    perorangan_check dk_surat_nikah ON dk_surat_nikah.bundle_hd_id = b_hd.id AND dk_surat_nikah.name = 'AKTA NIKAH/SURAT KETERANGAN NIKAH/SURAT CERAI'
                LEFT JOIN 
                    pt_check pt_ktp ON pt_ktp.bundle_hd_id = b_hd.id AND pt_ktp.name = 'KTP PENGURUS'
                LEFT JOIN 
                    pt_check pt_ad_art ON pt_ad_art.bundle_hd_id = b_hd.id AND pt_ad_art.name = 'AD/ART'
                LEFT JOIN 
                    pt_check pt_p_ad_art ON pt_p_ad_art.bundle_hd_id = b_hd.id AND pt_p_ad_art.name = 'PERUBAHAN AD/ART'
                LEFT JOIN 
                    waris_check waris_s_kematian ON waris_s_kematian.bundle_hd_id = b_hd.id AND waris_s_kematian.name = 'SURAT KEMATIAN'
                LEFT JOIN 
                    waris_check waris_a_kematian ON waris_a_kematian.bundle_hd_id = b_hd.id AND waris_a_kematian.name = 'AKTA KEMATIAN'
                LEFT JOIN 
                    waris_check waris_ket_ahli ON waris_ket_ahli.bundle_hd_id = b_hd.id AND waris_ket_ahli.name = 'SURAT KETERANGAN AHLI WARIS'
                LEFT JOIN 
                    waris_check waris_kuasa ON waris_kuasa.bundle_hd_id = b_hd.id AND waris_kuasa.name = 'SURAT KUASA AHLI WARIS'
                LEFT JOIN 
                    waris_check waris_ktp ON waris_ktp.bundle_hd_id = b_hd.id AND waris_ktp.name = 'KTP AHLI WARIS'
                LEFT JOIN 
                    waris_check waris_kk ON waris_kk.bundle_hd_id = b_hd.id AND waris_kk.name = 'KK AHLI WARIS'
                WHERE 
                    (hd.kategori_penjual = 'Perorangan' AND EXISTS (SELECT 1 FROM perorangan_check WHERE bundle_hd_id = b_hd.id AND meta_data IS NOT NULL))
                    OR (hd.kategori_penjual = 'PT' AND EXISTS (SELECT 1 FROM pt_check WHERE bundle_hd_id = b_hd.id AND meta_data IS NOT NULL))
                    OR (hd.kategori_penjual = 'Waris' AND EXISTS (SELECT 1 FROM waris_check WHERE bundle_hd_id = b_hd.id AND meta_data IS NOT NULL))
            )
            SELECT manager_id, manager_name, desa_name, nama_penjual, alashak, luas_m2, luas_ha, ktp_penjual, ktp_pasangan, kk, surat_nikah, ktp_pengurus,
            ad_art, perubahan_ad_art,
            CASE WHEN surat_kematian = 'V' AND akta_kematian = 'V' THEN 'V' ELSE '' END AS surat_akta_kematian, surat_ket_ahli_waris, surat_kuasa_waris,
            ktp_seluruh_waris, kk_seluruh_waris
            FROM main_query
            WHERE ktp_penjual = 'V' OR ktp_pasangan = 'V' OR kk = 'V' OR surat_nikah = 'V' OR ktp_pengurus = 'V' OR ad_art = 'V'
            OR perubahan_ad_art = 'V' OR (surat_kematian = 'V' AND akta_kematian = 'V') OR surat_ket_ahli_waris = 'V' 
            OR surat_kuasa_waris = 'V' OR ktp_seluruh_waris = 'V' OR kk_seluruh_waris = 'V'
            ORDER BY manager_id;
    """

    db_session = db.session
    response = await db_session.execute(query)
    results = response.all()

    datas = [KekuranganBerkasManagerSch(
        manager_id=row_data[0], manager_name=row_data[1], desa_name=row_data[2],
        nama_penjual=row_data[3], alashak=row_data[4], luas_m2=row_data[5], luas_ha=row_data[6],
        ktp_penjual=row_data[7], ktp_pasangan=row_data[8], kk=row_data[9], surat_nikah=row_data[10],
        ktp_pengurus=row_data[11], ad_art=row_data[12], perubahan_ad_art=row_data[13], surat_akta_kematian=row_data[14],
        surat_ket_ahli_waris=row_data[15], surat_kuasa_waris=row_data[16], ktp_seluruh_Waris=row_data[17], kk_seluruh_waris=row_data[18]
    ) for row_data in results]

    manager_ids = list(set(item.manager_id for item in datas))

    x = 4
    for manager_id in manager_ids:
        kekurangan_berkas_managers = list(map(lambda item: item, filter(lambda x: x.manager_id == manager_id, datas)))
        manager_name = next((manager.manager_name for manager in kekurangan_berkas_managers), None)
        nomor_urut = 1
        x +=1
        ws.cell(row=x, column=1, value=f"GROUP {manager_name.upper()}").font = Font(bold=True)
        for berkas_kurang in kekurangan_berkas_managers:
            x +=1
            ws.cell(row=x, column=1, value=nomor_urut)
            ws.cell(row=x, column=2, value=berkas_kurang.desa_name)
            ws.cell(row=x, column=3, value=berkas_kurang.nama_penjual)
            ws.cell(row=x, column=4, value=berkas_kurang.alashak)
            ws.cell(row=x, column=5, value=berkas_kurang.luas_m2)
            ws.cell(row=x, column=6, value=berkas_kurang.luas_ha)
            ws.cell(row=x, column=7, value=berkas_kurang.ktp_penjual).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=8, value=berkas_kurang.ktp_pasangan).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=9, value=berkas_kurang.kk).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=10, value=berkas_kurang.surat_nikah).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=11, value=berkas_kurang.ktp_pengurus).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=12, value=berkas_kurang.ad_art).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=13, value=berkas_kurang.perubahan_ad_art).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=14, value=berkas_kurang.surat_akta_kematian).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=15, value=berkas_kurang.surat_ket_ahli_waris).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=16, value=berkas_kurang.surat_kuasa_waris).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=17, value=berkas_kurang.ktp_seluruh_Waris).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=x, column=18, value=berkas_kurang.kk_seluruh_waris).alignment = Alignment(horizontal='center', vertical='center')
            nomor_urut += 1
        
        x +=1
        ws.cell(row=x, column=4, value=f"SUBTOTAL GROUP {manager_name.upper()}").font = Font(bold=True)
        ws.cell(row=x, column=5, value=f"{sum([bk.luas_m2 for bk in kekurangan_berkas_managers])}").font = Font(bold=True)
        ws.cell(row=x, column=6, value=f"{sum([bk.luas_ha for bk in kekurangan_berkas_managers])}").font = Font(bold=True)
        x +=1

    x +=1
    ws.cell(row=x, column=4, value=f"GRAND TOTAL KEKURANGAN IDENTITAS").font = Font(bold=True)
    ws.cell(row=x, column=5, value=f"{sum([bk.luas_m2 for bk in datas])}").font = Font(bold=True)
    ws.cell(row=x, column=6, value=f"{sum([bk.luas_ha for bk in datas])}").font = Font(bold=True)

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    
    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=kekurangan_berkas_manager.xlsx"})

@router.get("/kjb-lunas")
async def report_kjb_lunas(start_date:date | None = None, end_date:date|None = None):

    wb = Workbook()
    ws = wb.active

    ws.title =  "STATUS KJB-LUNAS"
    ws.firstHeader

    r1c5 = ws.cell(row=1, column=5)
    r1c5.font = Font(bold=True)
    r1c5.alignment = Alignment(horizontal='center', vertical='center')
    r1c5.value = "ADMIN MARKETING"
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)

    r1c7 = ws.cell(row=1, column=7)
    r1c7.font = Font(bold=True)
    r1c7.alignment = Alignment(horizontal='center', vertical='center')
    r1c7.value = "ANALIS"
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)

    r1c9 = ws.cell(row=1, column=9)
    r1c9.font = Font(bold=True)
    r1c9.alignment = Alignment(horizontal='center', vertical='center')
    r1c9.value = "PEMBAYARAN"
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=13)

    header_string = ["NO KJB", "TGL", "MARKETING", "LUAS", "TOTAL BERKAS MASUK", "TOTAL BERKAS BELUM MASUK", 
                    "SUDAH PETLOK", "BELUM PETLOK", "UTJ", "DP", "PELUNASAN", "LUNAS", "SISA"]
    

    for idx, val in enumerate(header_string):
        ws.cell(row=2, column=idx + 1, value=val).font = Font(bold=True)
        ws.cell(row=2, column=idx + 1, value=val).alignment = Alignment(horizontal='center', vertical='center')

    query_start_date = ""
    if start_date and end_date:
        query_start_date = "WHERE tanggal_kjb >=" f"'{start_date}'"
        query_start_date += " AND tanggal_kjb <" f"'{end_date}'"

    query = f"""
            WITH total_berkas_masuk AS (
            SELECT
                hd.id,
                SUM(dt.luas_surat) AS luas
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            WHERE 
                EXISTS (SELECT 1 FROM tanda_terima_notaris_hd WHERE kjb_dt_id = dt.id)
            GROUP BY
                hd.id
            ),
            total_berkas_belum_masuk AS (
            SELECT
                hd.id,
                SUM(dt.luas_surat) AS luas
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            WHERE 
                NOT EXISTS (SELECT 1 FROM tanda_terima_notaris_hd WHERE kjb_dt_id = dt.id)
            GROUP BY
                hd.id
            ),
            sudah_petlok AS (
            SELECT
                hd.id,
                SUM(dt.luas_surat) AS luas
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            WHERE 
                EXISTS (SELECT 1 FROM hasil_peta_lokasi WHERE kjb_dt_id = dt.id)
            GROUP BY
                hd.id
            ),
            belum_petlok AS (
            SELECT
                hd.id,
                SUM(dt.luas_surat) AS luas
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            WHERE 
                NOT EXISTS (SELECT 1 FROM hasil_peta_lokasi WHERE kjb_dt_id = dt.id)
            GROUP BY
                hd.id
            ),
            pembayaran AS (
            SELECT
                hd.id,
                hd.code,
                tr.jenis_bayar,
                SUM(dt.luas_surat) AS luas
            FROM
                kjb_hd hd
            CROSS JOIN
                (VALUES ('UTJ'), ('DP'), ('LUNAS'), ('PELUNASAN')) AS jenis_bayar(name)
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            INNER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            INNER JOIN
                invoice inv ON inv.bidang_id = hpl.bidang_id
            INNER JOIN
                termin tr ON tr.id = inv.termin_id AND tr.jenis_bayar = jenis_bayar.name
            WHERE
                inv.is_void != True
            GROUP BY
                hd.id, tr.jenis_bayar
            ),
            invoices AS (
            SELECT
                hd.id,
                SUM(COALESCE(inv.amount, 0)) as amount
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            INNER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            INNER JOIN
                bidang b ON b.id = hpl.bidang_id
            INNER JOIN
                invoice inv ON inv.bidang_id = b.id
            WHERE
                inv.is_void != True
                AND EXISTS (SELECT 1 FROM payment_detail pdt WHERE pdt.invoice_id = inv.id AND pdt.is_void != True)
            GROUP BY hd.id
            ),
            invoices_utj AS(
            SELECT
                hd.id,
                SUM(COALESCE(inv.amount, 0)) as amount
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            INNER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            INNER JOIN
                bidang b ON b.id = hpl.bidang_id
            INNER JOIN
                invoice inv ON inv.bidang_id = b.id
            INNER JOIN
                termin tr ON tr.id = inv.termin_id AND tr.jenis_bayar = 'UTJ'
            WHERE
                inv.is_void != True
            GROUP BY hd.id
            ),
            komponen_biaya AS (
            SELECT
                hd.id,
                SUM(COALESCE(kb.paid_amount, 0)) as amount
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            INNER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            INNER JOIN
                bidang_komponen_biaya kb ON kb.bidang_id = hpl.bidang_id
            WHERE
                kb.is_paid = True
                AND kb.is_void != True
                AND is_retur != True
                AND beban_pembeli != True
                AND tanggal_bayar IS NOT NULL
            GROUP BY hd.id
            ),
            total_harga AS (
            SELECT
                hd.id,
                SUM(COALESCE(b.harga_transaksi, 0) * COALESCE(b.luas_bayar, 0)) as total
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            INNER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            INNER JOIN
                bidang b ON b.id = hpl.bidang_id
            GROUP BY hd.id
            ),
            main_query AS (
            SELECT
                hd.id,
                hd.code,
                hd.tanggal_kjb,
                m.name AS manager_name,
                ROUND(SUM(COALESCE(dt.luas_surat, 0))::numeric/10000, 2) AS luas,
                ROUND(COALESCE(tbm.luas, 0)::numeric/10000, 2) AS total_berkas_masuk,
                ROUND(COALESCE(tbbm.luas, 0)::numeric/10000, 2) AS total_berkas_belum_masuk,
                ROUND(COALESCE(sp.luas, 0)::numeric/10000, 2) AS sudah_petlok,
                ROUND(COALESCE(bp.luas, 0)::numeric/10000, 2) AS belum_petlok,
                ROUND(COALESCE(utj.luas, 0)::numeric/10000, 2) AS utj,
                ROUND(COALESCE(dp.luas, 0)::numeric/10000, 2) AS dp,
                ROUND(COALESCE(pelunasan.luas, 0)::numeric/10000, 2) AS pelunasan,
                ROUND(COALESCE(lunas.luas, 0)::numeric/10000, 2) AS lunas,
                ROUND(COALESCE(
                        COALESCE(th.total, 0) - (COALESCE(inv.amount, 0) + COALESCE(inv_utj.amount, 0) + COALESCE(kb.amount, 0))
                , 0), 2) AS sisa
            FROM 
                kjb_hd hd
            INNER JOIN 
                kjb_dt dt ON hd.id = dt.kjb_hd_id
            LEFT OUTER JOIN
                hasil_peta_lokasi hpl ON hpl.kjb_dt_id = dt.id
            LEFT OUTER JOIN
                bidang b ON b.id = hpl.bidang_id
            LEFT OUTER JOIN
                manager m ON m.id = hd.manager_id
            LEFT OUTER JOIN
                total_berkas_masuk tbm ON tbm.id = hd.id
            LEFT OUTER JOIN
                total_berkas_belum_masuk tbbm ON tbbm.id = hd.id
            LEFT OUTER JOIN
                sudah_petlok sp ON sp.id = hd.id
            LEFT OUTER JOIN
                belum_petlok bp ON bp.id = hd.id
            LEFT OUTER JOIN
                pembayaran utj ON utj.id = hd.id AND utj.jenis_bayar = 'UTJ'
            LEFT OUTER JOIN
                pembayaran dp ON dp.id = hd.id AND dp.jenis_bayar = 'DP'
            LEFT OUTER JOIN
                pembayaran pelunasan ON pelunasan.id = hd.id AND pelunasan.jenis_bayar = 'PELUNASAN'
            LEFT OUTER JOIN
                pembayaran lunas ON lunas.id = hd.id AND lunas.jenis_bayar = 'LUNAS'
            LEFT OUTER JOIN
                invoices inv ON inv.id = hd.id
            LEFT OUTER JOIN
                invoices_utj inv_utj ON inv_utj.id = hd.id
            LEFT OUTER JOIN
                komponen_biaya kb ON kb.id = hd.id
            LEFT OUTER JOIN
                total_harga th ON th.id = hd.id
            GROUP BY
                hd.id, m.id, tbm.luas, tbbm.luas, sp.luas, bp.luas, utj.luas, dp.luas, pelunasan.luas, lunas.luas, 
                th.total, inv.amount, inv_utj.amount, kb.amount
            ORDER BY hd.code
            )
            SELECT * FROM main_query
            {query_start_date}
    """

    db_session = db.session
    response = await db_session.execute(query)
    results = response.all()

    x = 2
    for row_data in results:
        x += 1
        ws.cell(row=x, column=1, value=row_data[1])
        obj_tanggal_transaksi = datetime.strptime(str(row_data[2]), "%Y-%m-%d %H:%M:%S")
        tanggal_transaksi = obj_tanggal_transaksi.strftime("%d/%m/%Y")
        ws.cell(row=x, column=2, value=tanggal_transaksi)
        ws.cell(row=x, column=3, value=row_data[3])
        ws.cell(row=x, column=4, value=row_data[4])
        ws.cell(row=x, column=5, value=row_data[5])
        ws.cell(row=x, column=6, value=row_data[6])
        ws.cell(row=x, column=7, value=row_data[7])
        ws.cell(row=x, column=8, value=row_data[8])
        ws.cell(row=x, column=9, value=row_data[9])
        ws.cell(row=x, column=10, value=row_data[10])
        ws.cell(row=x, column=11, value=row_data[11])
        ws.cell(row=x, column=12, value=row_data[12])
        ws.cell(row=x, column=13, value=row_data[13])

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    
    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=kekurangan_berkas_manager.xlsx"})