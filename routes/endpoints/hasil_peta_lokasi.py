from uuid import UUID
from fastapi import APIRouter, status, Depends, UploadFile, Response, Request, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from fastapi_pagination import Params, Page
from fastapi_async_sqlalchemy import db
from sqlmodel import select, and_, or_
import crud
from models.hasil_peta_lokasi_model import HasilPetaLokasi, HasilPetaLokasiDetail
from models.worker_model import Worker
from models.bidang_model import Bidang
from models.bidang_overlap_model import BidangOverlap
from models.checklist_kelengkapan_dokumen_model import ChecklistKelengkapanDokumenHd, ChecklistKelengkapanDokumenDt
from schemas.hasil_peta_lokasi_sch import (HasilPetaLokasiSch, HasilPetaLokasiTaskUpdate, 
                                           HasilPetaLokasiCreateExtSch, HasilPetaLokasiByIdSch, 
                                           HasilPetaLokasiUpdateSch, HasilPetaLokasiUpdateExtSch,
                                           HasilPetaLokasiReadySpkExtSch, HasilPetaLokasiRemoveLink)
from schemas.hasil_peta_lokasi_detail_sch import (HasilPetaLokasiDetailCreateSch, HasilPetaLokasiDetailTaskUpdate)
from schemas.bidang_overlap_sch import BidangOverlapSch
from schemas.bidang_sch import BidangSch, BidangUpdateSch, BidangSrcSch
from schemas.bundle_hd_sch import BundleHdUpdateSch
from schemas.bundle_dt_sch import BundleDtCreateSch, BundleDtUpdateSch
from schemas.invoice_sch import InvoiceCreateSch, InvoiceUpdateSch
from schemas.payment_detail_sch import PaymentDetailCreateSch
from schemas.utj_khusus_detail_sch import UtjKhususDetailUpdateSch
from schemas.bidang_origin_sch import BidangOriginSch
from schemas.response_sch import (GetResponseBaseSch, GetResponsePaginatedSch, 
                                  PostResponseBaseSch, PutResponseBaseSch, create_response)
from common.exceptions import (IdNotFoundException, ContentNoChangeException, DocumentFileNotFoundException)
from common.generator import generate_code, CodeCounterEnum, generate_code_month
from common.enum import (StatusHasilPetaLokasiEnum, StatusBidangEnum, JenisBayarEnum, StatusPembebasanEnum,
                         JenisBidangEnum, HasilAnalisaPetaLokasiEnum, StatusLuasOverlapEnum, TipeOverlapEnum)
from services.gcloud_storage_service import GCStorageService
from services.gcloud_task_service import GCloudTaskService
from services.geom_service import GeomService
from services.helper_service import HelperService, KomponenBiayaHelper, BundleHelper
from services.history_service import HistoryService
from services.hasil_peta_lokasi_service import HasilPetaLokasiService
from shapely import wkb, to_wkt, wkt
from decimal import Decimal
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import geopandas as gpd
import pandas as pd
import roman



router = APIRouter()

@router.get("", response_model=GetResponsePaginatedSch[HasilPetaLokasiSch])
async def get_list(
            params: Params=Depends(), 
            order_by:str = None, 
            keyword:str = None, 
            filter_query:str=None,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    objs = await crud.hasil_peta_lokasi.get_multi_paginate_ordered_with_keyword_dict(params=params, order_by=order_by, keyword=keyword, filter_query=filter_query)
    return create_response(data=objs)

@router.get("/{id}", response_model=GetResponseBaseSch[HasilPetaLokasiByIdSch])
async def get_by_id(id:UUID):

    """Get an object by id"""

    obj = await crud.hasil_peta_lokasi.get_by_id(id=id)
    if obj and (obj.is_done or obj.is_done is None):
        return create_response(data=obj)
    elif obj and obj.is_done == False:
        raise HTTPException(status_code=422, detail="""Hasil Peta Lokasi ini masih terkunci 
                            dikarenakan sedang dalam proses pemotongan geometry bidang. Harap bersabar sebentar lagi.""")
    else:
        raise IdNotFoundException(HasilPetaLokasi, id)

@router.put("/upload-dokumen/{id}", response_model=PutResponseBaseSch[HasilPetaLokasiSch])
async def upload_dokumen(
            id:UUID, 
            background_task:BackgroundTasks,
            file: UploadFile = None,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Update a obj by its id"""

    obj_current = await crud.hasil_peta_lokasi.get_by_id(id=id)
    if not obj_current:
        raise IdNotFoundException(HasilPetaLokasi, id)

    if file:
        file_path = await GCStorageService().upload_file_dokumen(file=file, file_name=f'Hasil_Peta_Lokasi-{id}')
        object_updated = HasilPetaLokasiUpdateSch(file_path=file_path)
    
    obj_updated = await crud.hasil_peta_lokasi.update(obj_current=obj_current, obj_new=object_updated, updated_by_id=current_worker.id)
    obj_updated = await crud.hasil_peta_lokasi.get_by_id(id=obj_updated.id)

    background_task.add_task(BundleHelper().merge_hasil_lokasi, obj_updated.kjb_dt.bundle_hd_id, current_worker.id, obj_updated.id)

    return create_response(data=obj_updated)

@router.get("/download-file/{id}")
async def download_file(id:UUID):
    """Download File Dokumen"""

    obj_current = await crud.hasil_peta_lokasi.get_by_id(id=id)
    if not obj_current:
        raise IdNotFoundException(HasilPetaLokasi, id)
    if obj_current.file_path is None:
        raise DocumentFileNotFoundException(dokumenname=obj_current.alashak_kjb_dt)
    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=obj_current.file_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.alashak_kjb_dt)
    
    ext = obj_current.file_path.split('.')[-1]

    # return FileResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={obj_current.id}.{ext}"})
    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename=Hasil Peta Lokasi-{id}-{obj_current.id_bidang}.{ext}"
    return response

@router.get("/search/bidang", response_model=GetResponsePaginatedSch[BidangSrcSch])
async def get_list(
                params: Params=Depends(),
                keyword:str = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    status_ = [StatusBidangEnum.Belum_Bebas]
    query = select(Bidang.id, Bidang.id_bidang, Bidang.id_bidang_lama).select_from(Bidang
                    ).where(and_(
                                Bidang.status.in_(status_),
                                Bidang.jenis_bidang != JenisBidangEnum.Bintang,
                                Bidang.hasil_peta_lokasi == None
                            ))
    
    if keyword:
        query = query.filter(or_(
                                    Bidang.id_bidang.ilike(f'%{keyword}%'),
                                    Bidang.id_bidang_lama.ilike(f'%{keyword}%'),
                                )
                            )


    objs = await crud.bidang.get_multi_paginated(params=params, query=query)
    return create_response(data=objs)

@router.post("/create", response_model=PostResponseBaseSch[HasilPetaLokasiSch], status_code=status.HTTP_201_CREATED)
async def create(
            request: Request,
            sch: HasilPetaLokasiCreateExtSch,
            background_task:BackgroundTasks,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Create a new object"""

    db_session = db.session

    draft = None
    if sch.draft_id:
        draft = await crud.draft.get_by_id(id=sch.draft_id)
    
    for dt in sch.hasilpetalokasidetails:
        if dt.bidang_id is None:
            continue

        if dt.tipe_overlap == TipeOverlapEnum.BintangBatal and dt.status_luas != StatusLuasOverlapEnum.Menambah_Luas:
            raise HTTPException(status_code=422, detail=f"Apabila Bintang batal pada overlap, maka status luas harus Menambah Luas. Agar perhitungan luas bintang (DAMAI, BATAL, SISA BINTANG) sesuai")
        
        if dt.tipe_overlap == TipeOverlapEnum.BintangLanjut and dt.status_luas != StatusLuasOverlapEnum.Tidak_Menambah_Luas:
            raise HTTPException(status_code=422, detail=f"Apabila Bintang lanjut pada overlap, maka status luas harus Tidak Menambah Luas. Agar perhitungan luas bintang (DAMAI, BATAL, SISA BINTANG) sesuai")

    obj_current = await crud.hasil_peta_lokasi.get_by_kjb_dt_id(kjb_dt_id=sch.kjb_dt_id)
    if obj_current:
        raise ContentNoChangeException(detail="Alashak Sudah input hasil peta lokasi")

    sch.hasil_analisa_peta_lokasi = HasilAnalisaPetaLokasiEnum.Clear

    if len([dt for dt in sch.hasilpetalokasidetails if dt.bidang_id is not None]) > 0:
        sch.hasil_analisa_peta_lokasi = HasilAnalisaPetaLokasiEnum.Overlap
    
    sch.is_done = False if sch.bidang_id else True

    new_obj = await crud.hasil_peta_lokasi.create(obj_in=sch, created_by_id=current_worker.id, db_session=db_session, with_commit=False)
    
    if sch.bidang_id:
        bidang_current = await crud.bidang.get_by_id(id=sch.bidang_id)
        if bidang_current.geom :
            if isinstance(bidang_current.geom, str):
                pass
            else:
                bidang_current.geom = wkt.dumps(wkb.loads(bidang_current.geom.data, hex=True))
        if bidang_current.geom_ori :
            if isinstance(bidang_current.geom_ori, str):
                pass
            else:
                bidang_current.geom_ori = wkt.dumps(wkb.loads(bidang_current.geom_ori.data, hex=True))
        
        #add to bidang origin
        new_bidang_origin = BidangOriginSch.from_orm(bidang_current)
        await crud.bidang_origin.create_(obj_in=new_bidang_origin, db_session=db_session, with_commit=False)

        bidang_geom_updated = BidangSch(**sch.dict(), geom=wkt.dumps(wkb.loads(draft.geom.data, hex=True)))
        bidang_geom_updated.status_pembebasan = StatusPembebasanEnum.INPUT_PETA_LOKASI
        await crud.bidang.update(obj_current=bidang_current, obj_new=bidang_geom_updated, db_session=db_session, with_commit=False)

        details = [HasilPetaLokasiDetailTaskUpdate(tipe_overlap=x.tipe_overlap,
                                                bidang_id=str(x.bidang_id) if x.bidang_id is not None else x.bidang_id,
                                                luas_overlap=str(x.luas_overlap) if x.luas_overlap is not None else x.luas_overlap,
                                                keterangan=x.keterangan,
                                                draft_detail_id=str(x.draft_detail_id) if x.draft_detail_id is not None else x.draft_detail_id,
                                                status_luas=x.status_luas) 
                for x in sch.hasilpetalokasidetails]

        payload = HasilPetaLokasiTaskUpdate(bidang_id=str(new_obj.bidang_id) if new_obj.bidang_id is not None else new_obj.bidang_id,
                                                hasil_peta_lokasi_id=str(new_obj.id) if new_obj.id is not None else new_obj.id,
                                                kjb_dt_id=str(new_obj.kjb_dt_id) if new_obj.kjb_dt_id is not None else new_obj.kjb_dt_id,
                                                draft_id=str(sch.draft_id) if sch.draft_id is not None else sch.draft_id,
                                                from_updated=False,
                                                details=details)
        
        await db_session.commit()
        await db_session.refresh(new_obj)

        url1 = f'{request.base_url}landrope/hasilpetalokasi/task/insert-detail'
        GCloudTaskService().create_task(payload=payload.dict(), base_url=url1)

        url2 = f'{request.base_url}landrope/hasilpetalokasi/task/update-bidang'
        GCloudTaskService().create_task(payload=payload.dict(), base_url=url2)
    
    else:

        await db_session.commit()
        await db_session.refresh(new_obj)

    kjb_dt_current = await crud.kjb_dt.get(id=new_obj.kjb_dt_id)
    background_task.add_task(BundleHelper().merge_hasil_lokasi, kjb_dt_current.bundle_hd_id, current_worker.id, new_obj.id)

    new_obj = await crud.hasil_peta_lokasi.get_by_id(id=new_obj.id)

    return create_response(data=new_obj)

@router.put("/{id}", response_model=PutResponseBaseSch[HasilPetaLokasiSch])
async def update(
            id:UUID, 
            request:Request,
            sch:HasilPetaLokasiUpdateExtSch,
            background_task:BackgroundTasks,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Update a obj by its id"""

    db_session = db.session

    if sch.draft_id:
        draft = await crud.draft.get_by_id(id=sch.draft_id)

    for dt in sch.hasilpetalokasidetails:
        if dt.bidang_id is None:
            continue

        # bidang_overlap = await crud.bidang.get(id=dt.bidang_id)
        # if dt.luas_overlap > sch.luas_ukur:
        #     raise HTTPException(status_code=422, detail=f"Luas overlap {bidang_overlap.id_bidang} tidak boleh lebih besar dari luas ukur bidang yang menimpa")
        
        # if dt.luas_overlap > bidang_overlap.luas_surat:
        #     raise HTTPException(status_code=422, detail=f"Luas overlap {bidang_overlap.id_bidang} tidak boleh lebih besar dari luas suratnya {bidang_overlap.luas_surat}")
        
        if dt.tipe_overlap == TipeOverlapEnum.BintangBatal and dt.status_luas != StatusLuasOverlapEnum.Menambah_Luas:
            raise HTTPException(status_code=422, detail=f"Apabila Bintang batal pada overlap, maka status luas harus Menambah Luas. Agar perhitungan luas bintang (DAMAI, BATAL, SISA BINTANG) sesuai")
        
        if dt.tipe_overlap == TipeOverlapEnum.BintangLanjut and dt.status_luas != StatusLuasOverlapEnum.Tidak_Menambah_Luas:
            raise HTTPException(status_code=422, detail=f"Apabila Bintang lanjut pada overlap, maka status luas harus Tidak Menambah Luas. Agar perhitungan luas bintang (DAMAI, BATAL, SISA BINTANG) sesuai")

    obj_current = await crud.hasil_peta_lokasi.get_by_id(id=id)

    if not obj_current:
        raise IdNotFoundException(HasilPetaLokasi, id)
    
    #add history
    await HistoryService().create_history_hasil_peta_lokasi(obj_current=obj_current, worker_id=current_worker.id, db_session=db_session)
    
    #remove link bundle dan kelengkapan dokumen jika pada update yg dipilih bidang berbeda
    if obj_current.bidang_id != sch.bidang_id:
        if obj_current.bidang_id is not None:
            url = f'{request.base_url}landrope/hasilpetalokasi/task/remove-link-bidang-and-kelengkapan'
            payload = {"bidang_id" : str(obj_current.bidang_id), "worker_id" : str(obj_current.updated_by_id)}
            GCloudTaskService().create_task(payload=payload, base_url=url)
    
    sch.hasil_analisa_peta_lokasi = HasilAnalisaPetaLokasiEnum.Clear

    if len([dt for dt in sch.hasilpetalokasidetails if dt.bidang_id is not None]) > 0:
        sch.hasil_analisa_peta_lokasi = HasilAnalisaPetaLokasiEnum.Overlap

    #update hasil peta lokasi
    sch_updated = HasilPetaLokasiUpdateSch(**sch.dict())
    sch_updated.file_path = obj_current.file_path
    sch_updated.is_done = False if sch_updated.bidang_id else True
    obj_updated = await crud.hasil_peta_lokasi.update(obj_current=obj_current, obj_new=sch_updated,
                                                       updated_by_id=current_worker.id, db_session=db_session, with_commit=False)
    
    if sch.bidang_id:
        bidang_current = await crud.bidang.get_by_id(id=sch.bidang_id)
        if bidang_current.geom :
            if isinstance(bidang_current.geom, str):
                pass
            else:
                bidang_current.geom = wkt.dumps(wkb.loads(bidang_current.geom.data, hex=True))
        if bidang_current.geom_ori :
            if isinstance(bidang_current.geom_ori, str):
                pass
            else:
                bidang_current.geom_ori = wkt.dumps(wkb.loads(bidang_current.geom_ori.data, hex=True))

        bidang_geom_updated = BidangSch(**sch.dict(), geom=wkt.dumps(wkb.loads(draft.geom.data, hex=True))) 
        await crud.bidang.update(obj_current=bidang_current, obj_new=bidang_geom_updated, db_session=db_session, with_commit=False)

        details = [HasilPetaLokasiDetailTaskUpdate(tipe_overlap=x.tipe_overlap,
                                                bidang_id=str(x.bidang_id) if x.bidang_id is not None else x.bidang_id,
                                                luas_overlap=str(x.luas_overlap) if x.luas_overlap is not None else x.bidang_id,
                                                keterangan=x.keterangan,
                                                draft_detail_id=str(x.draft_detail_id) if x.draft_detail_id is not None else x.draft_detail_id,
                                                status_luas=x.status_luas) 
                for x in sch.hasilpetalokasidetails]

        payload = HasilPetaLokasiTaskUpdate(bidang_id=str(obj_updated.bidang_id) if obj_updated.bidang_id is not None else obj_updated.bidang_id,
                                                hasil_peta_lokasi_id=str(obj_updated.id) if obj_updated.id is not None else obj_updated.id,
                                                kjb_dt_id=str(obj_updated.kjb_dt_id) if obj_updated.kjb_dt_id is not None else obj_updated.kjb_dt_id,
                                                draft_id=str(sch.draft_id) if sch.draft_id is not None else sch.draft_id,
                                                from_updated=True,
                                                details=details)

        await db_session.commit()
        await db_session.refresh(obj_updated)

        url1 = f'{request.base_url}landrope/hasilpetalokasi/task/insert-detail'
        GCloudTaskService().create_task(payload=payload.dict(), base_url=url1)

        url2 = f'{request.base_url}landrope/hasilpetalokasi/task/update-bidang'
        GCloudTaskService().create_task(payload=payload.dict(), base_url=url2)

    else:
        await db_session.commit()
        await db_session.refresh(obj_updated)
    
    kjb_dt_current = await crud.kjb_dt.get(id=obj_updated.kjb_dt_id)
    background_task.add_task(BundleHelper().merge_hasil_lokasi, kjb_dt_current.bundle_hd_id, current_worker.id, obj_updated.id)

    obj_updated = await crud.hasil_peta_lokasi.get_by_id(id=obj_updated.id)

    return create_response(data=obj_updated)

@router.post("/task/insert-detail")
async def insert_detail(payload:HasilPetaLokasiTaskUpdate):
    
    db_session = db.session
    

    if payload.from_updated:
        db_session_remove = db.session
        hasil_peta_lokasi_current = await crud.hasil_peta_lokasi.get_by_id(id=payload.hasil_peta_lokasi_id)
        # kalau dia update, merge dulu semua geom hasil irisan di table bidang overlap dengan geom curent bidang bintang yg terkena overlap
        # agar geom current bintangnya kembali seperti sebelum terpotong
        # dengan kondisi yang tipe overlapnya bintang batal dan status luasnya menambah luas
        await merge_geom_kulit_bintang_with_geom_irisan_overlap(hasil_peta_lokasi_id=payload.hasil_peta_lokasi_id, worker_id=hasil_peta_lokasi_current.updated_by_id)

        # setelah itu hapus existing data hasil peta lokasi detail dan bidang overlap
        #remove existing data detail dan overlap
        list_overlap = [ov.bidang_overlap for ov in hasil_peta_lokasi_current.details if ov.bidang_overlap != None]

        await crud.hasil_peta_lokasi_detail.remove_multiple_data(list_obj=hasil_peta_lokasi_current.details, db_session=db_session_remove)
        await crud.bidangoverlap.remove_multiple_data(list_obj=list_overlap, db_session=db_session_remove)

        await db_session_remove.commit()

    hasil_peta_lokasi_current = await crud.hasil_peta_lokasi.get_by_id(id=payload.hasil_peta_lokasi_id)

    #bidang override
    bidang_override_current = await crud.bidang.get(id=hasil_peta_lokasi_current.bidang_id)
    override_geom_current = wkt.dumps(wkb.loads(bidang_override_current.geom.data, hex=True))
    gs_1 = gpd.GeoSeries.from_wkt([override_geom_current])
    gdf_1 = gpd.GeoDataFrame(geometry=gs_1)

    for dt in payload.details:
        bidang_overlap_id = None 
        if dt.draft_detail_id is not None:
            #input bidang overlap dari hasil analisa
            draft_detail = await crud.draft_detail.get(id=dt.draft_detail_id)
            if draft_detail is None:
                raise ContentNoChangeException(detail="Bidang Overlap tidak exists di Draft Detail")
            
            #Memotong geom kulit bintang apabila beririsan dengan kulit bintang, dengan kondisi:
            #- status peta lokasi LANJUT
            #- hasil peta lokasi detail MENAMBAH LUAS
            #- bidang yg tertimpa adalah KULIT BINTANG
            bidang_intersects_current = await crud.bidang.get_by_id(id=dt.bidang_id)
            if bidang_intersects_current.geom :
                bidang_intersects_current.geom = wkt.dumps(wkb.loads(bidang_intersects_current.geom.data, hex=True))
            if bidang_intersects_current.geom_ori :
                bidang_intersects_current.geom_ori = wkt.dumps(wkb.loads(bidang_intersects_current.geom_ori.data, hex=True))

            if hasil_peta_lokasi_current.status_hasil_peta_lokasi == StatusHasilPetaLokasiEnum.Lanjut and dt.status_luas == StatusLuasOverlapEnum.Menambah_Luas and bidang_intersects_current.jenis_bidang == JenisBidangEnum.Bintang:
                #1. Langkah pertama copy geom ke geom_ori pada data kulit bintang
                #jika kulit bintang sudah memiiki geom_ori maka lewati proses copy
                geom_ori = None
                if bidang_intersects_current.geom_ori is None:
                    geom_ori = bidang_intersects_current.geom

                #2. Langkah kedua bandingkan geom bidang hasil petlok dengan kulit bintang, kemudian jadikan hasil geom yang tidak tertiban menjadi geom baru
                gs_2 = gpd.GeoSeries.from_wkt([bidang_intersects_current.geom])
                gdf_2 = gpd.GeoDataFrame(geometry=gs_2)

                clipped_gdf = gdf_2.difference(gdf_1.unary_union)

                print(clipped_gdf.geometry)

                geom_new = None
                # Pengecekan geometri kosong
                if not clipped_gdf.is_empty.all():
                    # Buffer(0) dan convex hull
                    # clipped_gdf = clipped_gdf.buffer(0).convex_hull

                    # Pengecekan validitas geometri
                    is_result_valid = clipped_gdf.is_valid.all()
                    if not is_result_valid:
                        # Memperbaiki dengan buffer(0)
                        clipped_gdf = clipped_gdf.buffer(0)

                    # Pemilihan geometri yang benar
                    geom_new = GeomService.single_geometry_to_wkt(clipped_gdf.iloc[0])

                
                obj_new = BidangSch(geom_ori=(geom_ori if bidang_intersects_current.geom_ori is None else bidang_intersects_current.geom_ori), geom=(geom_new if geom_new is not None else bidang_intersects_current.geom))

                await crud.bidang.update(obj_current=bidang_intersects_current,
                                        obj_new=obj_new,
                                        db_session=db_session,
                                        with_commit=False)
            
            code = await generate_code(entity=CodeCounterEnum.BidangOverlap, db_session=db_session, with_commit=False)
            bidang_overlap_sch = BidangOverlapSch(
                                    code=code,
                                    parent_bidang_id=UUID(payload.bidang_id),
                                    parent_bidang_intersect_id=UUID(dt.bidang_id),
                                    luas=Decimal(dt.luas_overlap),
                                    status_luas=dt.status_luas,
                                    geom=wkt.dumps(wkb.loads(draft_detail.geom.data, hex=True)))
            
            new_obj_bidang_overlap = await crud.bidangoverlap.create(obj_in=bidang_overlap_sch, 
                                                                     db_session=db_session, 
                                                                     with_commit=False, 
                                                                     created_by_id=hasil_peta_lokasi_current.created_by_id)
            bidang_overlap_id = new_obj_bidang_overlap.id
            
        
        #input detail hasil peta lokasi
        detail_sch = HasilPetaLokasiDetailCreateSch(tipe_overlap=dt.tipe_overlap, 
                                                    bidang_id=UUID(dt.bidang_id) if dt.bidang_id is not None else dt.bidang_id, 
                                                    luas_overlap=Decimal(dt.luas_overlap),
                                                    keterangan=dt.keterangan, status_luas=dt.status_luas)
        
        detail_sch.hasil_peta_lokasi_id=hasil_peta_lokasi_current.id
        detail_sch.bidang_overlap_id=bidang_overlap_id

        await crud.hasil_peta_lokasi_detail.create(obj_in=detail_sch,
                                                   created_by_id=hasil_peta_lokasi_current.created_by_id, 
                                                   db_session=db_session, 
                                                   with_commit=False)

    await db_session.commit()
    await db_session.refresh(hasil_peta_lokasi_current)

    hasil_peta_lokasi_current_ = await crud.hasil_peta_lokasi.get(id=hasil_peta_lokasi_current.id)

    hasil_peta_lokasi_update = {"is_done" : True}
    await crud.hasil_peta_lokasi.update(obj_current=hasil_peta_lokasi_current_, 
                                        obj_new=hasil_peta_lokasi_update)

    return {"message":"successfully"} 

@router.post("/task/update-bidang")
async def update_bidang_override(payload:HasilPetaLokasiTaskUpdate, background_task:BackgroundTasks, request:Request):

    """Task update data bidang from hasil peta lokasi"""
    # try:
    db_session = db.session

    hasil_peta_lokasi = await crud.hasil_peta_lokasi.get_by_id_for_cloud(id=payload.hasil_peta_lokasi_id)
    kjb_dt_current = await crud.kjb_dt.get_by_id_for_cloud(id=payload.kjb_dt_id)
    kjb_hd_current = await crud.kjb_hd.get_by_id_for_cloud(id=kjb_dt_current.kjb_hd_id)
    tanda_terima_notaris_current = await crud.tandaterimanotaris_hd.get_one_by_kjb_dt_id(kjb_dt_id=kjb_dt_current.id)

    jenis_bidang = JenisBidangEnum.Standard
    status_bidang = StatusBidangEnum.Deal

    if hasil_peta_lokasi.status_hasil_peta_lokasi == StatusHasilPetaLokasiEnum.Batal:
        status_bidang = StatusBidangEnum.Batal

    if hasil_peta_lokasi.hasil_analisa_peta_lokasi == HasilAnalisaPetaLokasiEnum.Overlap:
        jenis_bidang = JenisBidangEnum.Overlap
    
    bidang_current = await crud.bidang.get_by_id(id=payload.bidang_id)
    bundle_hd_id_bidang = bidang_current.bundle_hd_id
    if bidang_current.geom :
        bidang_current.geom = wkt.dumps(wkb.loads(bidang_current.geom.data, hex=True))
    
    if bidang_current.geom_ori :
        bidang_current.geom_ori = wkt.dumps(wkb.loads(bidang_current.geom_ori.data, hex=True))
    
    if bidang_current.status == StatusBidangEnum.Bebas:
        status_bidang = bidang_current.status

    bidang_updated = BidangSch(
        jenis_bidang=jenis_bidang,
        status=status_bidang,
        group=kjb_dt_current.group,
        jenis_alashak=kjb_dt_current.jenis_alashak,
        jenis_surat_id=kjb_dt_current.jenis_surat_id,
        alashak=kjb_dt_current.alashak,
        manager_id=kjb_hd_current.manager_id,
        sales_id=kjb_hd_current.sales_id,
        pemilik_id=kjb_dt_current.pemilik_id,
        mediator=kjb_hd_current.mediator,
        telepon_mediator=kjb_hd_current.telepon_mediator,
        notaris_id=tanda_terima_notaris_current.notaris_id,
        bundle_hd_id=kjb_dt_current.bundle_hd_id,
        harga_akta=kjb_dt_current.harga_akta,
        harga_transaksi=kjb_dt_current.harga_transaksi,
        # harga_ptsl=kjb_dt_current.harga_ptsl,
        # is_ptsl=kjb_dt_current.is_ptsl
        )
    
    await crud.bidang.update(obj_current=bidang_current, 
                            obj_new=bidang_updated, 
                            updated_by_id=hasil_peta_lokasi.updated_by_id,
                            db_session=db_session,
                            with_commit=False)
    
    #jika kjb_dt memiliki bundle yang berbeda dengan bidang (case bidang dan bundle naik duluan diimport) sekalian merging apa yg ada di bundle kjb ke bundle bidang
    if kjb_dt_current.bundle_hd_id != bundle_hd_id_bidang:
        bundle_dts_bidang = await crud.bundledt.get_by_bundle_hd_id_for_merging(bundle_hd_id=bundle_hd_id_bidang)
        for bundle_dt_bidang in bundle_dts_bidang:
            bundle_dt_on_kjb_dt = await crud.bundledt.get_by_bundle_hd_id_and_dokumen_id(bundle_hd_id=kjb_dt_current.bundle_hd_id, dokumen_id=bundle_dt_bidang.dokumen_id)
            bundle_dt_on_kjb_dt_updated = BundleDtUpdateSch.from_orm(bundle_dt_on_kjb_dt)
            bundle_dt_on_kjb_dt_updated.meta_data = bundle_dt_bidang.meta_data
            bundle_dt_on_kjb_dt_updated.file_path = bundle_dt_bidang.file_path
            bundle_dt_on_kjb_dt_updated.riwayat_data = bundle_dt_bidang.riwayat_data
            bundle_dt_on_kjb_dt_updated.multiple_count = bundle_dt_bidang.multiple_count

            await crud.bundledt.update(obj_current=bundle_dt_on_kjb_dt, obj_new=bundle_dt_on_kjb_dt_updated, updated_by_id=hasil_peta_lokasi.updated_by_id, db_session=db_session)
    
        
    #bundle update planing
    bundle_current = await crud.bundlehd.get(id=kjb_dt_current.bundle_hd_id)
    bundle_updated = BundleHdUpdateSch.from_orm(bundle_current)
    bundle_updated.planing_id = hasil_peta_lokasi.planing_id
    await crud.bundlehd.update(obj_current=bundle_current, obj_new=bundle_updated, db_session=db_session, with_commit=False)
    
    # jika kjb_dt memiliki utj khusus
    bidang_ids = []
    utj_khusus_detail = await crud.utj_khusus_detail.get_by_kjb_dt_id(kjb_dt_id=payload.kjb_dt_id)
    if utj_khusus_detail:
        if utj_khusus_detail.invoice == None:
            today = date.today()
            month = roman.toRoman(today.month)
            year = today.year
            last_number_invoice = await generate_code_month(entity=CodeCounterEnum.Invoice, with_commit=False, db_session=db_session)
            invoice_sch = InvoiceCreateSch(
                        bidang_id=payload.bidang_id,
                        code=f"INV/{last_number_invoice}/{JenisBayarEnum.UTJ_KHUSUS}/LA/{month}/{year}",
                        amount=utj_khusus_detail.amount,
                        termin_id=utj_khusus_detail.utj_khusus.termin_id,
                        is_void=False
                    )
            
            invoice = await crud.invoice.create(obj_in=invoice_sch, db_session=db_session, with_commit=False, created_by_id=utj_khusus_detail.created_by_id)
            
            bidang_ids.append(invoice.bidang_id)

            #add payment detail
            payment_detail_sch = PaymentDetailCreateSch(payment_id=utj_khusus_detail.utj_khusus.payment_id, 
                                                        invoice_id=invoice.id, 
                                                        amount=invoice.amount, is_void=False)
            await crud.payment_detail.create(obj_in=payment_detail_sch, created_by_id=utj_khusus_detail.created_by_id, db_session=db_session, with_commit=False)

            utj_khusus_detail_updated = UtjKhususDetailUpdateSch(**utj_khusus_detail.dict(exclude={"invoice_id", "created_at", "updated_at"}), invoice_id=invoice.id)
            await crud.utj_khusus_detail.update(obj_current=utj_khusus_detail, obj_new=utj_khusus_detail_updated, db_session=db_session, with_commit=False)
        else:
            if utj_khusus_detail.invoice.bidang_id != payload.bidang_id:
                invoice_updated = InvoiceUpdateSch(bidang_id=payload.bidang_id)
                await crud.invoice.update(obj_current=utj_khusus_detail.invoice, obj_new=invoice_updated, db_session=db_session, with_commit=False)
    
    await db_session.commit()

    url3 = f'{request.base_url}landrope/hasilpetalokasi/task/generate-kelengkapan'
    GCloudTaskService().create_task(payload=payload.dict(), base_url=url3)

    background_task.add_task(HelperService().bidang_update_status, bidang_ids)
    background_task.add_task(KomponenBiayaHelper().calculated_all_komponen_biaya, [bidang_current.id])

    return {"message":"successfully"} 

@router.post("/task/generate-kelengkapan")
async def generate_kelengkapan_bidang_override(payload:HasilPetaLokasiTaskUpdate):

    """Task generate checklist kelengkapan dokumen from hasil peta lokasi"""

    db_session = db.session

    hasil_peta_lokasi = await crud.hasil_peta_lokasi.get_by_id_for_cloud(id=payload.hasil_peta_lokasi_id)
    kjb_dt_current = await crud.kjb_dt.get_by_id_for_cloud(id=payload.kjb_dt_id)
    kjb_hd_current = await crud.kjb_hd.get_by_id_for_cloud(id=kjb_dt_current.kjb_hd_id)

    bidang_current = await crud.bidang.get_by_id(id=payload.bidang_id)
    if bidang_current.bundle_hd_id is None:
        raise HTTPException(status_code=422, detail="bidang belum punya bundle")
    
    if bidang_current.geom :
        if isinstance(bidang_current.geom, str):
            pass
        else:
            bidang_current.geom = wkt.dumps(wkb.loads(bidang_current.geom.data, hex=True))
    if bidang_current.geom_ori :
        if isinstance(bidang_current.geom_ori, str):
            pass
        else:
            bidang_current.geom_ori = wkt.dumps(wkb.loads(bidang_current.geom_ori.data, hex=True))

    #generate kelengkapan dokumen
    if hasil_peta_lokasi.status_hasil_peta_lokasi != StatusHasilPetaLokasiEnum.Batal:
        checklist_kelengkapan_dokumen_hd_current = await crud.checklist_kelengkapan_dokumen_hd.get_by_bidang_id(bidang_id=payload.bidang_id)
        if checklist_kelengkapan_dokumen_hd_current is None:
            # removed_data = []
            # removed_data.append(checklist_kelengkapan_dokumen_hd_current)
            # await crud.checklist_kelengkapan_dokumen_hd.remove_multiple_data(list_obj=removed_data, db_session=db_session)

            master_checklist_dokumens = await crud.checklistdokumen.get_multi_by_jenis_alashak_and_kategori_penjual(
                jenis_alashak=kjb_dt_current.jenis_alashak,
                kategori_penjual=kjb_hd_current.kategori_penjual)
            
            checklist_kelengkapan_dts = []
            for master in master_checklist_dokumens:
                bundle_dt_current = await crud.bundledt.get_by_bundle_hd_id_and_dokumen_id_for_cloud(bundle_hd_id=bidang_current.bundle_hd_id, dokumen_id=master.dokumen_id)
                if not bundle_dt_current:
                    code = bidang_current.bundlehd.code + master.dokumen.code
                    bundle_dt_current = BundleDtCreateSch(code=code, 
                                                dokumen_id=master.dokumen_id,
                                                bundle_hd_id=bidang_current.bundle_hd_id)
                    
                    bundle_dt_current = await crud.bundledt.create(obj_in=bundle_dt_current, db_session=db_session, with_commit=False)

                checklist_kelengkapan_dt = ChecklistKelengkapanDokumenDt(
                    jenis_bayar=master.jenis_bayar,
                    dokumen_id=master.dokumen_id,
                    bundle_dt_id=bundle_dt_current.id,
                    created_by_id=hasil_peta_lokasi.updated_by_id,
                    updated_by_id=hasil_peta_lokasi.updated_by_id)
                
                checklist_kelengkapan_dts.append(checklist_kelengkapan_dt)
            
            checklist_kelengkapan_hd = ChecklistKelengkapanDokumenHd(bidang_id=payload.bidang_id, details=checklist_kelengkapan_dts)
            await crud.checklist_kelengkapan_dokumen_hd.create_and_generate(obj_in=checklist_kelengkapan_hd, 
                                                                            created_by_id=hasil_peta_lokasi.updated_by_id, 
                                                                            db_session=db_session, 
                                                                            with_commit=False)
            
            await db_session.commit()
            
    return {"message":"successfully"} 

@router.post("/task/remove-link-bidang-and-kelengkapan")
async def remove_link_bidang_and_kelengkapan(payload:HasilPetaLokasiRemoveLink):

    """Task Remove link bundle and remove existing kelengkapan dokumen"""

    db_session = db.session

    #bidang
    #rollback from bidang origin when exists
    bidang_origin = await crud.bidang_origin.get(id=payload.bidang_id)
    if bidang_origin:
        bidang_old = await crud.bidang.get_by_id(id=payload.bidang_id)
        if bidang_origin.geom :
            bidang_origin.geom = wkt.dumps(wkb.loads(bidang_origin.geom.data, hex=True))

        if bidang_origin.geom_ori :
            bidang_origin.geom_ori = wkt.dumps(wkb.loads(bidang_origin.geom_ori.data, hex=True))

        if bidang_old.geom :
            bidang_old.geom = wkt.dumps(wkb.loads(bidang_old.geom.data, hex=True))

        if bidang_old.geom_ori :
            bidang_old.geom_ori = wkt.dumps(wkb.loads(bidang_old.geom_ori.data, hex=True))

        bidang_old_updated = BidangUpdateSch(**bidang_origin.dict())
        await crud.bidang.update(obj_current=bidang_old, obj_new=bidang_old_updated, db_session=db_session, with_commit=False, origin=True)
       
    else:
        bidang_old = await crud.bidang.get_by_id(id=payload.bidang_id)
        if bidang_old.geom :
            bidang_old.geom = wkt.dumps(wkb.loads(bidang_old.geom.data, hex=True))

        if bidang_old.geom_ori :
            bidang_old.geom_ori = wkt.dumps(wkb.loads(bidang_old.geom_ori.data, hex=True))
        
        bidang_old_updated = BidangUpdateSch(jenis_bidang=JenisBidangEnum.Standard,
                                    status=StatusBidangEnum.Belum_Bebas,
                                    group=None,
                                    jenis_surat_id=None,
                                    manager_id=None,
                                    sales_id=None,
                                    mediator=None,
                                    telepon_mediator=None,
                                    notaris_id=None,
                                    bundle_hd_id=None,
                                    harga_akta=None,
                                    harga_transaksi=None,
                                    status_pembebasan=None,
                                    luas_bayar=None,
                                    luas_clear=None,
                                    luas_gu_perorangan=None,
                                    luas_gu_pt=None,
                                    luas_nett=None,
                                    luas_pbt_perorangan=None,
                                    luas_pbt_pt=None,
                                    luas_produk=None,
                                    luas_proses=None,
                                    luas_ukur=None,
                                    #planing_id=None,
                                    #skpt_id=None
                                    )

        await crud.bidang.update(obj_current=bidang_old, obj_new=bidang_old_updated, db_session=db_session, with_commit=False)

    # kelengkapan dokumen
    checklist_kelengkapan_hd_old = await crud.checklist_kelengkapan_dokumen_hd.get_by_bidang_id(bidang_id=payload.bidang_id)
    if checklist_kelengkapan_hd_old:
        await crud.checklist_kelengkapan_dokumen_hd.remove(id=checklist_kelengkapan_hd_old.id, db_session=db_session, with_commit=False)
    if bidang_origin:
        await crud.bidang_origin.remove(id=bidang_origin.id, db_session=db_session, with_commit=False)

    await db_session.commit()

    return {"message" : "successfully remove link bidang and kelengkapan dokumen"}

async def merge_geom_kulit_bintang_with_geom_irisan_overlap(hasil_peta_lokasi_id:UUID, worker_id:UUID):
    
    bidang_overlaps = await crud.bidangoverlap.get_multi_kulit_bintang_batal_by_petlok_id(hasil_peta_lokasi_id=hasil_peta_lokasi_id)

    for ov in bidang_overlaps:
        db_session_update = db.session
        overlap = await crud.bidangoverlap.get(id=ov.id)
        if overlap.geom:
            overlap.geom = wkt.dumps(wkb.loads(overlap.geom.data, hex=True))
        
        ov_series = gpd.GeoSeries.from_wkt([overlap.geom])
        ov_gdf = gpd.GeoDataFrame(geometry=ov_series)

        bidang_bintang = await crud.bidang.get_by_id(id=overlap.parent_bidang_intersect_id)
        if bidang_bintang.geom:
            bidang_bintang.geom = wkt.dumps(wkb.loads(bidang_bintang.geom.data, hex=True))
        
        if bidang_bintang.geom_ori:
            bidang_bintang.geom_ori = wkt.dumps(wkb.loads(bidang_bintang.geom_ori.data, hex=True))

        bd_series = gpd.GeoSeries.from_wkt([bidang_bintang.geom])
        bd_series = bd_series.buffer(0).convex_hull

        # Anda juga dapat mencoba untuk memperbaiki dengan melakukan perbaikan secara manual
        # Misalnya, jika bd_series[0] adalah Polygon yang menyebabkan kesalahan, Anda dapat mencoba:
        polygon = bd_series[0]
        if not polygon.is_valid:
            corrected_polygon = polygon.buffer(0).convex_hull
            bd_series[0] = corrected_polygon

        gdf = gpd.GeoDataFrame(geometry=pd.concat([bd_series, ov_series], ignore_index=True))

        union_geom = gdf.geometry.unary_union
        
        # union_geom = bd_gdf.union(ov_gdf)
        if isinstance(union_geom, gpd.GeoSeries):
            union_geom = union_geom[0]

        # Cek apakah hasilnya tidak kosong
        if not union_geom.is_empty:
            geom_union = GeomService.single_geometry_to_wkt(union_geom)
            bidang_bintang_updated = {"geom" : geom_union}
            await crud.bidang.update(obj_current=bidang_bintang, obj_new=bidang_bintang_updated, db_session=db_session_update)
        else:
            print("Hasil gabungan geometri kosong atau tidak valid.")
        
        # union_geom = bd_gdf.unary_union(ov_gdf.unary_union)
        # geom_union = GeomService.single_geometry_to_wkt(union_geom[0])

        # bidang_bintang_updated = {"geom" : geom_union}

@router.get("/report/excel-detail")
async def report_detail(start_date:date | None = None, end_date:date|None = None, 
                        project_ids:str|None=None, desa_ids:str|None=None,
                        status:StatusHasilPetaLokasiEnum|None=None):

    wb = Workbook()
    ws = wb.active

    ws.title =  "REPORT HASIL PETA LOKASI DETAIL"
    ws.firstHeader

    ws.cell(row=1, column=2, value="LAPORAN HASIL PETA LOKASI DETAIL")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.cell(row=2, column=2, value=f"CUT OFF DATE {start_date} S/D {end_date}")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)

    header_string = ["ID Bidang", "Alashak", "Desa", "Project", "Nama Pemilik Tanah", "Nomor SK", "PT SK", "Status SK",
                    "Jenis Surat", "No Peta", "L Surat", "L Ukur", "L Nett", "L Clear", "L GU Perorangan", "L GU PT",
                    "L Proses", "L PBT Perorangan", "L PBT PT", "Status Peta Lokasi", "Hasil Analisa", "ID Bidang Overlap",
                    "Alashak Overlap", "Jenis Bidang", "Tipe Overlap", "L Overlap", "Status Luas", "Keterangan"]
    
    end_column = len(header_string)

    for idx, val in enumerate(header_string):
        ws.cell(row=3, column=idx + 1, value=val).font = Font(bold=True)
        if idx + 1 < 22:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        else:
            ws.cell(row=3, column=idx + 1, value=val).fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")

    query_start_date = ""
    if start_date and end_date:
        query_start_date = "WHERE date(hpl.created_at) >=" f"'{start_date}'"
        query_start_date += " AND date(hpl.created_at) <=" f"'{end_date}'"

    if project_ids:
        project_id_split = [f"'{str(req)}'" for req in project_ids.split(',')]
        projects = ",".join(project_id_split)

        if query_start_date == "":
            query_start_date = f"WHERE p.id IN ({projects})"
        else:
            query_start_date += f" AND p.id IN ({projects})"
    
    if desa_ids:
        desa_id_split = [f"'{str(req)}'" for req in desa_ids.split(',')]
        desas = ",".join(desa_id_split)

        if query_start_date == "":
            query_start_date = f"WHERE d.id IN ({desas})"
        else:
            query_start_date += f" AND d.id IN ({desas})"
    
    if status:
        if query_start_date == "":
            query_start_date = f"WHERE hpl.status_hasil_peta_lokasi = '{status}'"
        else:
            query_start_date += f" AND hpl.status_hasil_peta_lokasi = '{status}'"
    
    query = f"""
            SELECT
            bd.id_bidang,
            bd.alashak,
            d.name as desa_name,
            p.name as project_name,
            pm.name as pemilik_name,
            sk.nomor_sk,
            pt.name as ptsk_name,
            sk.status as status_sk,
            hpl.jenis_alashak,
            hpl.no_peta,
            hpl.luas_surat,
            hpl.luas_ukur,
            hpl.luas_nett,
            hpl.luas_clear,
            hpl.luas_gu_perorangan,
            hpl.luas_gu_pt,
            hpl.luas_proses,
            hpl.luas_pbt_perorangan,
            hpl.luas_pbt_pt,
            hpl.status_hasil_peta_lokasi,
            hpl.hasil_analisa_peta_lokasi,
            bd_dt.id_bidang as id_bidang_overlap,
            bd_dt.alashak as alashak_overlap,
            bd_dt.jenis_bidang as jenis_bidang_overlap,
            hpl_dt.tipe_overlap,
            hpl_dt.luas_overlap,
            hpl_dt.status_luas,
            hpl_dt.keterangan
            FROM hasil_peta_lokasi hpl
            INNER JOIN hasil_peta_lokasi_detail hpl_dt ON hpl.id = hpl_dt.hasil_peta_lokasi_id
            LEFT OUTER JOIN bidang bd ON bd.id = hpl.bidang_id
            LEFT OUTER JOIN bidang bd_dt ON bd_dt.id = hpl_dt.bidang_id
            LEFT OUTER JOIN planing pl ON pl.id = hpl.planing_id
            LEFT OUTER JOIN desa d ON d.id = pl.desa_id
            LEFT OUTER JOIN project p ON p.id = pl.project_id
            LEFT OUTER JOIN skpt sk ON sk.id = hpl.skpt_id
            LEFT OUTER JOIN ptsk pt ON pt.id = sk.ptsk_id
            LEFT OUTER JOIN pemilik pm ON pm.id = hpl.pemilik_id
            {query_start_date}
    """

    db_session = db.session
    response = await db_session.execute(query)
    result = response.all()

    x = 3
    for row_data in result:
        x += 1
        ws.cell(row=x, column=1, value=row_data[0])
        ws.cell(row=x, column=2, value=row_data[1])
        ws.cell(row=x, column=3, value=row_data[2])
        ws.cell(row=x, column=4, value=row_data[3])
        ws.cell(row=x, column=5, value=row_data[4])
        ws.cell(row=x, column=6, value=row_data[5])
        ws.cell(row=x, column=7, value=row_data[6])
        ws.cell(row=x, column=8, value=row_data[7])
        ws.cell(row=x, column=9, value=row_data[8])
        ws.cell(row=x, column=10, value=row_data[9])
        ws.cell(row=x, column=11, value=row_data[10])
        ws.cell(row=x, column=12, value=row_data[11])
        ws.cell(row=x, column=13, value=row_data[12])
        ws.cell(row=x, column=14, value=row_data[13])
        ws.cell(row=x, column=15, value=row_data[14])
        ws.cell(row=x, column=16, value=row_data[15])
        ws.cell(row=x, column=17, value=row_data[16])
        ws.cell(row=x, column=18, value=row_data[17])
        ws.cell(row=x, column=19, value=row_data[18])
        ws.cell(row=x, column=20, value=row_data[19])
        ws.cell(row=x, column=21, value=row_data[20])
        ws.cell(row=x, column=22, value=row_data[21])
        ws.cell(row=x, column=23, value=row_data[22])
        ws.cell(row=x, column=24, value=row_data[23])
        ws.cell(row=x, column=25, value=row_data[24])
        ws.cell(row=x, column=26, value=row_data[25])
        ws.cell(row=x, column=27, value=row_data[26])
        ws.cell(row=x, column=28, value=row_data[27])

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=generated_excel.xlsx"})

@router.get("/ready/spk", response_model=GetResponsePaginatedSch[HasilPetaLokasiReadySpkExtSch])
async def ready_spk(keyword:str | None = None, params: Params=Depends(), current_worker:Worker = Depends(crud.worker.get_active_worker)):

    result = await HasilPetaLokasiService().get_ready_spk(keyword=keyword, params=params)
    
    return create_response(data=result)

@router.get("/report/ready/spk/")
async def report_ready_spk(current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    return await HasilPetaLokasiService().get_report_ready_spk()

@router.get("/generate-kelengkapan/bidang")
async def generate_kelengkapan_bidang():

    db_session = db.session
    db_session1 = db.session
    query = """
            select h.id as hasil_peta_lokasi_id, h.bidang_id, h.kjb_dt_id from hasil_peta_lokasi h
            left outer join checklist_kelengkapan_dokumen_hd c on h.bidang_id = c.bidang_id
            inner join bidang b on b.id = h.bidang_id
            where h.bidang_id is not null and c.id is null and status_hasil_peta_lokasi != 'Batal' 
            and b.bundle_hd_id is not null
            """
    
    response = await db_session1.execute(query)
    result = response.fetchall()

    bidang_ids = []
    for res in result:
        bidang_ids.append({"hasil_peta_lokasi_id" : res[0], "bidang_id" : res[1], "kjb_dt_id" : res[2]})

    row = 0
    for payload in bidang_ids:
        if row == 10:
            break

        hasil_peta_lokasi = await crud.hasil_peta_lokasi.get_by_id_for_cloud(id=payload["hasil_peta_lokasi_id"])
        kjb_dt_current = await crud.kjb_dt.get_by_id_for_cloud(id=payload["kjb_dt_id"])
        kjb_hd_current = await crud.kjb_hd.get_by_id_for_cloud(id=kjb_dt_current.kjb_hd_id)

        bidang_current = await crud.bidang.get_by_id(id=payload["bidang_id"])
        if bidang_current.bundle_hd_id is None:
            raise HTTPException(status_code=422, detail="bidang belum punya bundle")
        
        if bidang_current.geom :
            if isinstance(bidang_current.geom, str):
                pass
            else:
                bidang_current.geom = wkt.dumps(wkb.loads(bidang_current.geom.data, hex=True))
        if bidang_current.geom_ori :
            if isinstance(bidang_current.geom_ori, str):
                pass
            else:
                bidang_current.geom_ori = wkt.dumps(wkb.loads(bidang_current.geom_ori.data, hex=True))

        #generate kelengkapan dokumen
        if hasil_peta_lokasi.status_hasil_peta_lokasi != StatusHasilPetaLokasiEnum.Batal:
            checklist_kelengkapan_dokumen_hd_current = await crud.checklist_kelengkapan_dokumen_hd.get_by_bidang_id(bidang_id=payload["bidang_id"])
            if checklist_kelengkapan_dokumen_hd_current:
                removed_data = []
                removed_data.append(checklist_kelengkapan_dokumen_hd_current)
                await crud.checklist_kelengkapan_dokumen_hd.remove_multiple_data(list_obj=removed_data, db_session=db_session)

            master_checklist_dokumens = await crud.checklistdokumen.get_multi_by_jenis_alashak_and_kategori_penjual(
                jenis_alashak=kjb_dt_current.jenis_alashak,
                kategori_penjual=kjb_hd_current.kategori_penjual)
            
            checklist_kelengkapan_dts = []
            for master in master_checklist_dokumens:
                bundle_dt_current = await crud.bundledt.get_by_bundle_hd_id_and_dokumen_id_for_cloud(bundle_hd_id=bidang_current.bundle_hd_id, dokumen_id=master.dokumen_id)
                if not bundle_dt_current:
                    code = bidang_current.bundlehd.code + master.dokumen.code
                    bundle_dt_current = BundleDtCreateSch(code=code, 
                                                dokumen_id=master.dokumen_id,
                                                bundle_hd_id=bidang_current.bundle_hd_id)
                    
                    bundle_dt_current = await crud.bundledt.create(obj_in=bundle_dt_current, db_session=db_session, with_commit=False)

                checklist_kelengkapan_dt = ChecklistKelengkapanDokumenDt(
                    jenis_bayar=master.jenis_bayar,
                    dokumen_id=master.dokumen_id,
                    bundle_dt_id=bundle_dt_current.id,
                    created_by_id=hasil_peta_lokasi.updated_by_id,
                    updated_by_id=hasil_peta_lokasi.updated_by_id)
                
                checklist_kelengkapan_dts.append(checklist_kelengkapan_dt)
            
            checklist_kelengkapan_hd = ChecklistKelengkapanDokumenHd(bidang_id=payload["bidang_id"], details=checklist_kelengkapan_dts)
            await crud.checklist_kelengkapan_dokumen_hd.create_and_generate(obj_in=checklist_kelengkapan_hd, 
                                                                            created_by_id=hasil_peta_lokasi.updated_by_id, 
                                                                            db_session=db_session, 
                                                                            with_commit=False)
            
            row += 1

    await db_session.commit()
    return {"message" : "successfully"}