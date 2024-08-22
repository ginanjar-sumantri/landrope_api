from uuid import UUID
from fastapi import APIRouter, status, Depends, HTTPException, Response, BackgroundTasks, UploadFile, Request
from fastapi.responses import StreamingResponse
from fastapi_pagination import Params
from fastapi_async_sqlalchemy import db
from sqlmodel import select, and_, or_, func, delete
from sqlalchemy import cast, Date
from sqlalchemy.orm import selectinload
from models import (Spk, Bidang, HasilPetaLokasi, KjbDt, KjbHd, Manager, Worker, 
                    Invoice, Termin, Planing, Workflow, WorkflowNextApprover, BidangKomponenBiaya)
from schemas.spk_sch import (SpkSch, SpkCreateSch, SpkUpdateSch, SpkByIdSch, SpkListSch, SpkVoidSch)
from schemas.bundle_dt_sch import BundleDtRiwayatSch
from schemas.bidang_sch import BidangSrcSch, BidangForSPKByIdExtSch
from schemas.workflow_sch import WorkflowSystemCreateSch, WorkflowSystemAttachmentSch, WorkflowUpdateSch
from schemas.response_sch import (PostResponseBaseSch, GetResponseBaseSch, DeleteResponseBaseSch, GetResponsePaginatedSch, PutResponseBaseSch, create_response)
from common.enum import (JenisBayarEnum,  WorkflowEntityEnum, WorkflowLastStatusEnum)
from common.exceptions import (IdNotFoundException, DocumentFileNotFoundException)
from services.history_service import HistoryService
from services.helper_service import KomponenBiayaHelper, BundleHelper
from services.workflow_service import WorkflowService
from services.gcloud_storage_service import GCStorageService
from services.spk_services import SpkService
from services.encrypt_service import encrypt_id
from datetime import date
from typing import Dict
import crud
import json
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter()

@router.post("/create", response_model=PostResponseBaseSch[SpkSch], status_code=status.HTTP_201_CREATED)
async def create(
            sch: SpkCreateSch,
            background_task:BackgroundTasks,
            request: Request,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Create a new object"""

    if sch.jenis_bayar in [JenisBayarEnum.LUNAS, JenisBayarEnum.PELUNASAN, JenisBayarEnum.PAJAK]:
        spk_exists = await crud.spk.get_by_bidang_id_jenis_bayar(bidang_id=sch.bidang_id, jenis_bayar=sch.jenis_bayar)
        if spk_exists:
            raise HTTPException(status_code=422, detail="SPK bidang dengan jenis bayar yang sama sudah ada")

    #Filter
    if (sch.is_draft or False) is False:
        if sch.jenis_bayar == JenisBayarEnum.BIAYA_LAIN:
            beban_biaya_ids = [x.beban_biaya_id for x in sch.spk_beban_biayas]
            await SpkService().filter_biaya_lain(beban_biaya_ids=beban_biaya_ids, bidang_id=sch.bidang_id)

        if sch.jenis_bayar in [JenisBayarEnum.DP, JenisBayarEnum.LUNAS, JenisBayarEnum.PELUNASAN]:
            await SpkService().filter_have_input_peta_lokasi(bidang_id=sch.bidang_id)

            #if bidang beginning balance lolosin aja untuk filter ini asal ada bundle spk sebelumnya
            spk_beginning_balance = await crud.spk.get_by_bidang_id_jenis_bayar(bidang_id=sch.bidang_id, jenis_bayar=JenisBayarEnum.BEGINNING_BALANCE)
            meta_data = await crud.bundledt.get_meta_data_by_dokumen_name_and_bidang_id(dokumen_name="SURAT PERINTAH KERJA", bidang_id=sch.bidang_id)

            if spk_beginning_balance is None:
                bundle_dt_ids = [dokumen.bundle_dt_id for dokumen in sch.spk_kelengkapan_dokumens]
                await SpkService().filter_kelengkapan_dokumen(bundle_dt_ids=bundle_dt_ids)
            elif spk_beginning_balance and meta_data is None:
                raise HTTPException(status_code=422, detail="Bidang memiliki beginning balance, hanya dokumen spk sebelumnya belum diupload dibundle")
            else:
                pass

    if sch.jenis_bayar in [JenisBayarEnum.DP, JenisBayarEnum.PELUNASAN]:
        await SpkService().filter_with_same_kjb_termin(bidang_id=sch.bidang_id, kjb_termin_id=sch.kjb_termin_id)
    #EndFilter

    new_obj = await SpkService().create_spk(sch=sch, current_worker=current_worker, request=request)
    new_obj = await crud.spk.get_by_id(id=new_obj.id)

    if (sch.is_draft or False) is False:
        bidang_ids = []
        bidang_ids.append(new_obj.bidang_id)

        background_task.add_task(KomponenBiayaHelper().calculated_all_komponen_biaya, bidang_ids)
    
    return create_response(data=new_obj)

@router.get("", response_model=GetResponsePaginatedSch[SpkListSch])
async def get_list(
                start_date: date | None = None,
                end_date: date | None = None,
                outstanding: bool | None = False,
                params: Params = Depends(), 
                order_by: str | None = None, 
                keyword: str | None = None, 
                filter_query: str | None = None,
                filter_list: str | None = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(Spk)
    query = query.join(Bidang, Spk.bidang_id == Bidang.id
                ).outerjoin(HasilPetaLokasi, HasilPetaLokasi.bidang_id == Bidang.id
                ).outerjoin(KjbDt, KjbDt.id == HasilPetaLokasi.kjb_dt_id
                ).outerjoin(KjbHd, KjbHd.id == KjbDt.kjb_hd_id
                ).outerjoin(Manager, Manager.id == Bidang.manager_id
                ).outerjoin(Workflow, and_(Workflow.reference_id == Spk.id, Workflow.entity == WorkflowEntityEnum.SPK))

    if filter_list == "list_approval":
        subquery_workflow = (select(Workflow.reference_id).join(Workflow.workflow_next_approvers
                                ).where(and_(WorkflowNextApprover.email == current_worker.email, 
                                            Workflow.entity == WorkflowEntityEnum.SPK,
                                            Workflow.last_status != WorkflowLastStatusEnum.COMPLETED))
                    ).distinct()
        
        query = query.filter(Spk.id.in_(subquery_workflow))
    
    if keyword:
        query = query.filter(
            or_(
                Spk.code.ilike(f'%{keyword}%'),
                Bidang.id_bidang.ilike(f'%{keyword}%'),
                Bidang.alashak.ilike(f'%{keyword}%'),
                Bidang.group.ilike(f'%{keyword}%'),
                KjbHd.code.ilike(f'%{keyword}%'),
                Manager.name.ilike(f'%{keyword}%'),
                Spk.jenis_bayar.ilike(f'%{keyword}%'),
                Workflow.last_status.ilike(f'%{keyword}%'),
                Workflow.step_name.ilike(f'%{keyword}%')
            )
        )

    if filter_query:
        filter_query = json.loads(filter_query)
        for key, value in filter_query.items():
                query = query.where(getattr(Spk, key) == value)
    
    if start_date and end_date:
        query = query.filter(cast(Spk.created_at, Date).between(start_date, end_date))
    
    if outstanding:
        subquery = (
            select(Invoice.spk_id)
            .join(Termin, Termin.id == Invoice.termin_id)
            .filter(Invoice.is_void != True, ~Termin.jenis_bayar.in_(['UTJ', 'UTJ_KHUSUS', 'BEGINNING_BALANCE']))
            .distinct()
        )
        query = query.filter(~Spk.jenis_bayar.in_(['BEGINNING_BALANCE', 'PAJAK']))
        query = query.filter(~Spk.id.in_(subquery))
        query = query.filter(Workflow.last_status == WorkflowLastStatusEnum.COMPLETED)


    query = query.distinct()

    objs = await crud.spk.get_multi_paginated_ordered(params=params, query=query)

    items = []
    reference_ids = [spk.id for spk in objs.data.items]
    workflows = await crud.workflow.get_by_reference_ids(reference_ids=reference_ids, entity=WorkflowEntityEnum.SPK)

    for obj in objs.data.items:
        workflow = next((wf for wf in workflows if wf.reference_id == obj.id), None)
        if workflow:
            obj.status_workflow = workflow.last_status
            obj.step_name_workflow = workflow.step_name

        items.append(obj)

    objs.data.items = items

    return create_response(data=objs)

@router.get("/export/excel")
async def get_report(
                start_date:date|None = None,
                end_date:date|None = None,
                outstanding:bool|None = False,
                keyword:str = None, 
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    filename:str = ''
    query = select(Spk).select_from(Spk
                    ).join(Bidang, Spk.bidang_id == Bidang.id)
    
    if keyword:
        query = query.filter(
            or_(
                Spk.code.ilike(f'%{keyword}%'),
                Bidang.id_bidang.ilike(f'%{keyword}%'),
                Bidang.alashak.ilike(f'%{keyword}%')
            )
        )
    
    if start_date and end_date:
        query = query.filter(cast(Spk.created_at, Date).between(start_date, end_date))
        filename = str(filename)

    if outstanding:
        subquery = (
            select(Invoice.spk_id)
            .join(Termin, Termin.id == Invoice.termin_id)
            .filter(Invoice.is_void != True, ~Termin.jenis_bayar.in_(['UTJ', 'UTJ_KHUSUS', 'BEGINNING_BALANCE']))
            .distinct()
        )
        query = query.filter(~Spk.jenis_bayar.in_(['BEGINNING_BALANCE', 'PAJAK']))
        query = query.filter(~Spk.id.in_(subquery))

    query = query.distinct()
    query = query.options(selectinload(Spk.bidang
                                ).options(selectinload(Bidang.pemilik)
                                ).options(selectinload(Bidang.planing
                                                ).options(selectinload(Planing.project)
                                                ).options(selectinload(Planing.desa)
                                                )
                                ).options(selectinload(Bidang.manager)
                                )
                    )

    objs = await crud.spk.get_multi_no_page(query=query)

    data = []
    
    reference_ids = [s.id for s in objs]
    workflows = await crud.workflow.get_by_reference_ids(reference_ids=reference_ids, entity=WorkflowEntityEnum.SPK)

    wb = Workbook()
    ws = wb.active

    header_string = ["Id Bidang", "Id Bidang Lama", "Group", "Pemilik", "Alashak", "Project", "Desa", "Luas Surat", 
                    "Jenis Bayar", "Manager", "Tanggal Buat", "Status Workflow", "Created By"]

    for idx, val in enumerate(header_string):
        ws.cell(row=1, column=idx + 1, value=val).font = Font(bold=True)

    x = 1
    for spk in objs:
        workflow = next((wf for wf in workflows if wf.reference_id == spk.id), None)
        status_workflow = "-"
        if workflow:
            status_workflow = workflow.last_status if workflow.last_status == WorkflowLastStatusEnum.COMPLETED else workflow.step_name or "-"

        x += 1
        ws.cell(row=x, column=1, value=spk.id_bidang)
        ws.cell(row=x, column=2, value=spk.id_bidang_lama or "")
        ws.cell(row=x, column=3, value=spk.group)
        ws.cell(row=x, column=4, value=spk.bidang.pemilik_name)
        ws.cell(row=x, column=5, value=spk.alashak)
        ws.cell(row=x, column=6, value=spk.bidang.project_name)
        ws.cell(row=x, column=7, value=spk.bidang.desa_name)
        ws.cell(row=x, column=8, value=spk.bidang.luas_surat).number_format = '0.00'
        ws.cell(row=x, column=9, value=spk.jenis_bayar)
        ws.cell(row=x, column=10, value=spk.manager_name)
        ws.cell(row=x, column=11, value=spk.created_at)
        ws.cell(row=x, column=12, value=status_workflow)
        ws.cell(row=x, column=13, value=spk.created_name)

    excel_data = BytesIO()

    # Simpan workbook ke objek BytesIO
    wb.save(excel_data)

    # Set posisi objek BytesIO ke awal
    excel_data.seek(0)
    

    return StreamingResponse(excel_data,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                             headers={"Content-Disposition": "attachment; filename=spk_data.xlsx"})
   
@router.get("/{id}", response_model=GetResponseBaseSch[SpkByIdSch])
async def get_by_id(id:UUID):

    """Get an object by id"""

    obj_return = await SpkService().get_by_id_spk(id=id)

    return create_response(data=obj_return)

@router.put("/{id}", response_model=PutResponseBaseSch[SpkSch])
async def update(id:UUID, 
                 sch:SpkUpdateSch,
                 background_task:BackgroundTasks,
                 request:Request,
                 current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Update a obj by its id"""

    obj_current = await crud.spk.get_by_id(id=id)

    if not obj_current:
        raise IdNotFoundException(Spk, id)
    
    if obj_current.is_void:
        raise HTTPException(status_code=422, detail="SPK sudah void")
    
    workflow = await crud.workflow.get_by_reference_id(reference_id=obj_current.id)
    
    #workflow
    if sch.jenis_bayar != JenisBayarEnum.PAJAK:
        if workflow:
            msg_error_wf = "SPK Approval Has Been Completed!" if WorkflowLastStatusEnum.COMPLETED else "SPK Approval Need Approval!"
            
            if sch.file is None:
                if workflow.last_status not in [WorkflowLastStatusEnum.NEED_DATA_UPDATE, WorkflowLastStatusEnum.REJECTED]:
                    raise HTTPException(status_code=422, detail=f"Failed update. Detail : {msg_error_wf}")
   
    #filter
    if (sch.is_draft or False) is False:
        bidang_current = await crud.bidang.get_by_id_for_spk(id=obj_current.bidang_id)

        if sch.jenis_bayar not in [JenisBayarEnum.BIAYA_LAIN]:
            if bidang_current.has_invoice_lunas:
                raise HTTPException(status_code=422, detail="Failed Update. Detail : Bidang already have Invoice Lunas")
        else:
            beban_biaya_ids = [x.beban_biaya_id for x in sch.spk_beban_biayas]
            await SpkService().filter_biaya_lain(beban_biaya_ids=beban_biaya_ids, bidang_id=sch.bidang_id)

        if sch.jenis_bayar in [JenisBayarEnum.LUNAS, JenisBayarEnum.PELUNASAN, JenisBayarEnum.PAJAK]:
            spk_exists = await crud.spk.get_by_bidang_id_jenis_bayar(bidang_id=sch.bidang_id, jenis_bayar=sch.jenis_bayar)
            if spk_exists:
                if spk_exists.id != obj_current.id:
                    raise HTTPException(status_code=422, detail="SPK bidang dengan jenis bayar yang sama sudah ada")

        if sch.jenis_bayar in [JenisBayarEnum.DP, JenisBayarEnum.LUNAS, JenisBayarEnum.PELUNASAN]:
            await SpkService().filter_have_input_peta_lokasi(bidang_id=sch.bidang_id)

            #if bidang beginning balance lolosin aja untuk filter ini asal ada bundle spk sebelumnya
            spk_beginning_balance = await crud.spk.get_by_bidang_id_jenis_bayar(bidang_id=sch.bidang_id, jenis_bayar=JenisBayarEnum.BEGINNING_BALANCE)
            meta_data = await crud.bundledt.get_meta_data_by_dokumen_name_and_bidang_id(dokumen_name="SURAT PERINTAH KERJA", bidang_id=sch.bidang_id)

            if spk_beginning_balance is None:
                bundle_dt_ids = [dokumen.bundle_dt_id for dokumen in sch.spk_kelengkapan_dokumens]
                await SpkService().filter_kelengkapan_dokumen(bundle_dt_ids=bundle_dt_ids)
            elif spk_beginning_balance and meta_data is None:
                raise HTTPException(status_code=422, detail="Bidang memiliki beginning balance, hanya dokumen spk sebelumnya belum diupload dibundle")
            else:
                pass
        
    if sch.jenis_bayar in [JenisBayarEnum.DP, JenisBayarEnum.PELUNASAN]:
        await SpkService().filter_with_same_kjb_termin(bidang_id=sch.bidang_id, kjb_termin_id=sch.kjb_termin_id, spk_id=obj_current.id)
    #end filter

    obj_updated = await SpkService().update_spk(sch=sch, obj_current=obj_current, bidang_id=sch.bidang_id, current_worker=current_worker, request=request)
    obj_updated = await crud.spk.get_by_id(id=obj_updated.id)

    if (sch.is_draft or False) is False:
        bidang_ids = []
        bidang_ids.append(bidang_current.id)

        background_task.add_task(KomponenBiayaHelper().calculated_all_komponen_biaya, bidang_ids)
        background_task.add_task(SpkService().generate_printout, obj_updated.id)

    return create_response(data=obj_updated)

@router.delete("/delete", response_model=DeleteResponseBaseSch[SpkSch], status_code=status.HTTP_200_OK)
async def delete_(id:UUID, current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Delete a object"""

    obj_current = await crud.spk.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Spk, id)
    
    obj_deleted = await crud.spk.remove(id=id)

    return obj_deleted

@router.get("/search/bidang", response_model=GetResponsePaginatedSch[BidangSrcSch])
async def search_bidang(
                params: Params=Depends(),
                keyword:str = None,
                current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """Gets a paginated list objects"""

    query = select(Bidang.id, Bidang.id_bidang, Bidang.alashak).select_from(Bidang
                    ).join(HasilPetaLokasi, Bidang.id == HasilPetaLokasi.bidang_id)
    
    if keyword:
        query = query.filter(or_(Bidang.id_bidang.ilike(f'%{keyword}%'), 
                                Bidang.alashak.ilike(f'%{keyword}%')))


    objs = await crud.bidang.get_multi_paginated(params=params, query=query)
    return create_response(data=objs)

@router.get("/search/bidang/{id}", response_model=GetResponseBaseSch[BidangForSPKByIdExtSch])
async def search_bidang_by_id(id:UUID):

    """Get an object by id"""

    obj_return = await SpkService().search_bidang_by_id(id=id)
    return create_response(data=obj_return)
    
@router.get("/riwayat/bundle_dt", response_model=GetResponseBaseSch[list[BundleDtRiwayatSch]])
async def get_riwayat_bundle_dt(bundle_dt_id:UUID):

    datas = await SpkService().get_riwayat_bundle_dt(bundle_dt_id=bundle_dt_id)

    return create_response(data=datas)
    
@router.get("/print-out/{id}")
async def printout(id:UUID | str, current_worker:Worker = Depends(crud.worker.get_active_worker)):

    """Print out"""

    obj_current = await crud.spk.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Spk, id)
    
    file_path = await SpkService().generate_printout(id=id)
    
    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=file_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    
    ext = obj_current.file_path.split('.')[-1]

    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename={obj_current.code.replace('/', '_')}.{ext}"
    return response

@router.put("/void/{id}", response_model=GetResponseBaseSch[SpkByIdSch])
async def void(id:UUID, 
            sch:SpkVoidSch,
            background_task:BackgroundTasks,
            current_worker:Worker = Depends(crud.worker.get_active_worker)):
    
    """void a obj by its ids"""
    db_session = db.session

    obj_current = await crud.spk.get_by_id(id=id)

    if not obj_current:
        raise IdNotFoundException(Spk, id)
    
    if obj_current.has_invoice_active:
        raise HTTPException(status_code=422, detail="Failed void. Detail : Spk have invoice active in Memo Pembayaran!")
    
    #schema for history
    spk_history = await SpkService().get_by_id_spk(id=id)
    await HistoryService().create_history_spk(spk=spk_history, worker_id=current_worker.id, db_session=db_session)
    
    obj_updated = SpkUpdateSch.from_orm(obj_current)
    obj_updated.is_void = True
    obj_updated.void_reason = sch.void_reason
    obj_updated.void_by_id = current_worker.id
    obj_updated.void_at = date.today()

    obj_updated = await crud.spk.update(obj_current=obj_current, obj_new=obj_updated, updated_by_id=current_worker.id, db_session=db_session)
    obj_updated = await SpkService().get_by_id_spk(id=id)
    
    background_task.add_task(delete_all_bidang_komponen_biaya, obj_current.bidang_id)

    return create_response(data=obj_updated) 

async def delete_all_bidang_komponen_biaya(bidang_id:UUID):

    """Delete All Bidang Komponen Biaya When all SPK Void"""
    objs_spk_active = await crud.spk.get_multi_spk_active_by_bidang_id(bidang_id=bidang_id)
    if len(objs_spk_active) == 0:
        db_session = db.session
        await db_session.execute(delete(BidangKomponenBiaya).where(BidangKomponenBiaya.bidang_id == bidang_id))
        await db_session.commit()

@router.post("/task-workflow")
async def task_workflow(payload:Dict, request:Request):

    id = payload.get("id", None)
    additional_info = payload.get("additional_info", None)

    obj = await crud.spk.get_by_id(id=id)

    if not obj:
        raise IdNotFoundException(Spk, id)
    
    await SpkService().task_workflow(obj=obj, additional_info=additional_info, request=request)
    
    return {"message" : "successfully"}

@router.post("/task-upload-printout")
async def task_upload_printout(payload:Dict):

    id = payload.get("id", None)

    obj = await crud.spk.get_by_id(id=id)

    if not obj:
        raise IdNotFoundException(Spk, id)
    
    await SpkService().task_generate_printout_and_merge_to_bundle(obj=obj)
    
    return {"message" : "successfully"}

@router.get("/download-file/{id}")
async def download_file(id:UUID):
    """Download File Dokumen"""

    obj_current = await crud.spk.get(id=id)
    if not obj_current:
        raise IdNotFoundException(Spk, id)
    if obj_current.file_upload_path is None:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    try:
        file_bytes = await GCStorageService().download_dokumen(file_path=obj_current.file_upload_path)
    except Exception as e:
        raise DocumentFileNotFoundException(dokumenname=obj_current.code)
    
    ext = obj_current.file_path.split('.')[-1]

    # return FileResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={obj_current.id}.{ext}"})
    response = Response(content=file_bytes, media_type="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename=Hasil Peta Lokasi-{id}-{obj_current.code}.{ext}"
    return response