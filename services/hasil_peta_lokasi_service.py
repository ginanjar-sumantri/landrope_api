from fastapi.responses import Response
from fastapi_pagination import Page
from schemas.hasil_peta_lokasi_sch import HasilPetaLokasiReadySpkExtSch, JenisBayarOnReady
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import crud
import math

class HasilPetaLokasiService:

    # LIST BIDANG YANG SUDAH SIAP UNTUK DIBUAT SPKNYA BERDASARKAN JENIS BAYAR & KELENGKAPAN DOKUMENNYA
    # 
    async def get_ready_spk(self, keyword, params):

        spks = await crud.hasil_peta_lokasi.get_ready_spk(keyword=keyword)

        # AMBIL SEMUA ID
        ids = [data.id for data in spks]

        # DISTINCT ID YANG SAMA
        ids = list(set(ids))

        # SETUP UNTUK PAGINATION
        start = (params.page - 1) * params.size
        end = params.page * params.size
        total_items = len(ids)
        pages = math.ceil(total_items / params.size)

        # AMBIL ID YANG AKAN DILOOPING DAN DIMAPING SCHEMA SEJUMLAH SIZE DAN PAGE YANG DIINGINKAN
        ids = ids[start:end]

        objs:list[HasilPetaLokasiReadySpkExtSch] = []

        # INISIALISASI SEMUA OBJECT SESUAI JUMLAH ID YANG TERPILIH PADA PAGINATION
        for id in ids:
            datas = list(filter(lambda obj: obj.id == id, spks))

            # GET FIRST DATA FOR DEFAULT
            data = datas[0]

            # SETUP DATA BIDANG
            data_ready_spk = HasilPetaLokasiReadySpkExtSch(id=data.id, id_bidang=data.id_bidang, alashak=data.alashak,
                                                        satuan_bayar=data.satuan_bayar, satuan_harga=data.satuan_harga, 
                                                        planing_id=data.planing_id, kjb_hd_code=data.kjb_hd_code, group=data.group, 
                                                        manager_id=data.manager_id, pemilik_id=data.pemilik_id, kjb_hd_id=data.kjb_hd_id,
                                                        hasil_petlok_id=data.hasil_petlok_id, req_petlok_id=data.req_petlok_id, 
                                                        kjb_dt_id=data.kjb_dt_id, project_name=data.project_name, desa_name=data.desa_name, 
                                                        manager_name=data.manager_name)
                
            

            # CEK APAKAH PEMILIK BIDANG MEMILIKI REKENING
            rekenings = await crud.rekening.get_by_pemilik_id(pemilik_id=data.pemilik_id)
            data_ready_spk.rekening_on_ready = False if not rekenings else True

            # CEK APAKAH KJB HARGA SUDAH DITENTUKAN
            data_ready_spk.kjb_harga_on_ready = False if data.kjb_harga_id is None else True

            # CEK APAKAH SK PADA BIDANG SUDAH DITENTUKAN
            data_ready_spk.sk_on_ready = False if data.skpt_id is None else True

            # SETUP JENIS BAYAR ON READY SESUAI DOKUMEN KELENGKAPAN
            jenis_bayar_on_ready = []
            for jns in datas:
                j = JenisBayarOnReady(kjb_termin_id=jns.kjb_termin_id, nilai=jns.nilai, jenis_bayar=jns.jenis_bayar)
                jenis_bayar_on_ready.append(j)

            data_ready_spk.jenis_bayar_on_ready = jenis_bayar_on_ready

            objs.append(data_ready_spk)
        

        result = Page(items=objs, size=params.size, page=params.page, pages=pages, total=total_items)

        return result
    
    async def get_report_ready_spk(self):

        wb = Workbook()
        ws = wb.active

        spks = await crud.hasil_peta_lokasi.get_ready_spk()

        # AMBIL SEMUA ID
        ids = [data.id for data in spks]

        # DISTINCT ID YANG SAMA
        ids = list(set(ids))

        header_string = ["Id Bidang", 
                "Alashak",
                "Project", 
                "Desa",
                "No KJB",
                "Group",
                "Manager",
                "Jenis Bayar",
                "Kelengkapan Dokumen",
                "Rekening",
                "Termin KJB",
                "PTSK"]

        for idx, val in enumerate(header_string):
            ws.cell(row=1, column=idx + 1, value=val).font = Font(bold=True)

        x = 1
        for id in ids:
            datas = list(filter(lambda obj: obj.id == id, spks))

            # GET FIRST DATA FOR DEFAULT
            data = datas[0]

            # SETUP JENIS BAYAR ON READY SESUAI DOKUMEN KELENGKAPAN
            jenis_bayar_on_ready = [f"{jns.jenis_bayar}({jns.nilai or ''})" if jns.jenis_bayar in ['DP', 'PELUNASAN'] else f"{jns.jenis_bayar}" for jns in datas]
            jenis_bayar = ", ".join(jenis_bayar_on_ready)

            # CEK APAKAH PEMILIK BIDANG MEMILIKI REKENING
            rekenings = await crud.rekening.get_by_pemilik_id(pemilik_id=data.pemilik_id)
            rekening_on_ready = False if not rekenings else True

            # CEK APAKAH KJB HARGA SUDAH DITENTUKAN
            kjb_harga_on_ready = False if data.kjb_harga_id is None else True

            # CEK APAKAH SK PADA BIDANG SUDAH DITENTUKAN
            sk_on_ready = False if data.skpt_id is None else True

            x += 1
            ws.cell(row=x, column=1, value=data.id_bidang)
            ws.cell(row=x, column=2, value=data.alashak or "")
            ws.cell(row=x, column=3, value=data.project_name)
            ws.cell(row=x, column=4, value=data.desa_name)
            ws.cell(row=x, column=5, value=data.kjb_hd_code)
            ws.cell(row=x, column=6, value=data.group)
            ws.cell(row=x, column=7, value=data.manager_name)
            ws.cell(row=x, column=8, value=jenis_bayar)
            ws.cell(row=x, column=9, value='READY')
            ws.cell(row=x, column=10, value='READY' if rekening_on_ready else 'NOT SET')
            ws.cell(row=x, column=11, value='READY' if kjb_harga_on_ready else 'NOT SET')
            ws.cell(row=x, column=12, value='READY' if sk_on_ready else 'NOT SET')

        
        excel_data = BytesIO()

        # Simpan workbook ke objek BytesIO
        wb.save(excel_data)

        # Set posisi objek BytesIO ke awal
        excel_data.seek(0)

        response = Response(content=excel_data.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = f"attachment; filename=ready_spk.xlsx"
        return response
        




